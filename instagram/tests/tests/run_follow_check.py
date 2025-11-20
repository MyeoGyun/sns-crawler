"""
보조 스크립트: 이미 수집된 XLSX 파일을 기반으로 팔로우 여부 확인만 실행.
사용법: python instagram/tests/tests/run_follow_check.py
"""

import asyncio
import sys
import re
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import load_workbook
from playwright.async_api import async_playwright

BASE_DIR = Path(__file__).resolve().parents[2]
REPO_ROOT = BASE_DIR.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from instagram.comment_extractor import (  # noqa: E402
    load_cookies_from_env,
    load_cookies_to_context,
    parse_follow_target_input,
    check_follow_status_for_comments,
)
DEFAULT_WORKBOOK = BASE_DIR / "comments.xlsx"


def load_comments_from_excel(path: Path, limit: int | None = None) -> List[Dict[str, Any]]:
    """엑셀 파일에서 작성자/프로필 링크 및 기존 팔로우 여부 컬럼을 추출"""
    wb = load_workbook(path)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    follow_cols: Dict[int, str] = {}
    for idx, header in enumerate(header_row):
        if not header:
            continue
        match = re.match(r"팔로우여부\((.+)\)", str(header).strip())
        if match:
            follow_cols[idx] = match.group(1)

    comments: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        username = (row[0] or "" ).strip()
        profile_link = (row[1] or "" ).strip()
        if not username:
            continue
        comment: Dict[str, Any] = {"username": username, "profile_link": profile_link}
        follow_status: Dict[str, str] = {}
        for col_idx, label in follow_cols.items():
            value = row[col_idx] if col_idx < len(row) else None
            if value is None:
                continue
            value_str = str(value).strip()
            if value_str:
                follow_status[label] = value_str
        if follow_status:
            comment["follow_status"] = follow_status
        comments.append(comment)
        if limit and len(comments) >= limit:
            break
    return comments


async def run_follow_check(workbook_path: Path, targets: str, limit: int | None, headless: bool) -> None:
    """Playwright 세션을 열어 특정 키워드에 대한 팔로우 여부만 검사"""
    target_specs = parse_follow_target_input(targets)
    if not target_specs:
        print("⚠️  유효한 팔로우 대상이 없습니다. 쉼표로 구분된 계정/URL을 입력하세요.")
        return

    comments = load_comments_from_excel(workbook_path, limit)
    if not comments:
        print("⚠️  엑셀에서 유효한 댓글 데이터를 찾지 못했습니다.")
        return

    cookies = load_cookies_from_env()
    if not cookies:
        print("❌ .env에서 Instagram 쿠키를 읽지 못했습니다.")
        return

    print(f"\n📁 대상 파일: {workbook_path}")
    print(f"👥 검사 대상 사용자: {len({c['username'] for c in comments})}명")
    print(f"🎯 팔로우 검색 키워드: {', '.join([spec['label'] for spec in target_specs])}\n")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless)
        context = await browser.new_context()
        await load_cookies_to_context(context, cookies)

        page = await context.new_page()
        await page.goto("https://www.instagram.com/")
        await page.wait_for_timeout(2000)

        await check_follow_status_for_comments(page, comments, target_specs)

        print("\n📊 검사 결과")
        print("─" * 40)
        for comment in comments:
            statuses = comment.get("follow_status") or {}
            if not statuses:
                continue
            joined = ", ".join(f"{label}: {mark}" for label, mark in statuses.items())
            print(f"@{comment['username']:<20} → {joined}")

        await browser.close()


def prompt_int(prompt: str, default: int | None = None) -> int | None:
    value = input(prompt).strip()
    if not value:
        return default
    try:
        parsed = int(value)
        return parsed if parsed > 0 else default
    except ValueError:
        return default


async def main():
    default_label = str(DEFAULT_WORKBOOK.relative_to(REPO_ROOT))
    workbook_input = input(f"엑셀 경로 입력 (엔터={default_label}): ").strip()
    if not workbook_input:
        workbook_path = DEFAULT_WORKBOOK
    else:
        candidates = [
            Path(workbook_input).expanduser(),
            REPO_ROOT / workbook_input,
            BASE_DIR / workbook_input,
        ]
        workbook_path = next((path for path in candidates if path.exists()), candidates[0])
    if not workbook_path.exists():
        print(f"❌ 파일을 찾을 수 없습니다: {workbook_path}")
        print("    예) comments.xlsx, instagram/comments.xlsx 등으로 입력해 주세요.")
        return

    print("예시: handle123, handle123|표시이름, handle123=>표시이름")
    targets = input("확인할 계정/키워드 (쉼표 구분): ").strip()
    if not targets:
        print("❌ 최소 한 개 이상의 계정/키워드가 필요합니다.")
        return

    limit = prompt_int("검사할 최대 사용자 수 (엔터=전체): ", default=None)
    headless_input = input("브라우저 숨김 모드? (y/N): ").strip().lower()
    headless = headless_input in {"y", "yes"}

    await run_follow_check(workbook_path, targets, limit, headless)


if __name__ == "__main__":
    asyncio.run(main())
