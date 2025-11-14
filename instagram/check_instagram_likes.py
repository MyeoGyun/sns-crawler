"""
Instagram 게시글 좋아요 확인 및 엑셀 업데이트 스크립트
사용법: python3 check_instagram_likes.py
"""

import asyncio
import io
import openpyxl
import msoffcrypto
from playwright.async_api import async_playwright
import sys

# 설정
INSTAGRAM_POST_URL = 'https://www.instagram.com/lotte_xylitol.official/p/DP5dlatE-mA/'
INSTAGRAM_USERNAME = 'aud.virus_0209'
INSTAGRAM_PASSWORD = '@Audrbs1063710'
EXCEL_FILE_PATH = '/Users/dinggyun/dev/retweet_check/(잘함) 10월 자일리톨 수능 응원 이벤트 당첨자 리스트_1023.xlsx'
EXCEL_PASSWORD = 'xylitol_25'
OUTPUT_FILE_PATH = '/Users/dinggyun/dev/retweet_check/(잘함) 10월 자일리톨 수능 응원 이벤트 당첨자 리스트_1023_updated.xlsx'

async def get_instagram_likes(post_url):
    """Instagram 게시글의 좋아요 사용자 목록을 가져옵니다."""
    print(f'\n{"="*60}')
    print('Instagram 좋아요 목록 수집 시작')
    print(f'{"="*60}')
    print(f'\n게시글 URL: {post_url}')

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=['--force-dark-mode=0']
        )

        context = await browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
            viewport={'width': 1280, 'height': 800}
        )

        page = await context.new_page()

        try:
            # Instagram 로그인
            print('\nInstagram 로그인 중...')
            await page.goto('https://www.instagram.com/accounts/login/')
            await page.wait_for_timeout(3000)

            try:
                await page.wait_for_selector('input[name="username"]', timeout=10000)

                print('사용자명 입력 중...')
                await page.fill('input[name="username"]', INSTAGRAM_USERNAME)
                await page.wait_for_timeout(1000)

                print('비밀번호 입력 중...')
                await page.fill('input[name="password"]', INSTAGRAM_PASSWORD)
                await page.wait_for_timeout(1000)

                print('로그인 버튼 클릭...')
                await page.click('button[type="submit"]')
                await page.wait_for_timeout(5000)

                print('✓ 로그인 완료!')

                # "나중에 하기" 버튼 처리
                try:
                    not_now = page.locator('button:has-text("나중에 하기"), button:has-text("Not Now")')
                    if await not_now.count() > 0:
                        await not_now.first.click()
                        await page.wait_for_timeout(2000)
                except:
                    pass

                # 알림 버튼 처리
                try:
                    not_now = page.locator('button:has-text("나중에 하기"), button:has-text("Not Now")')
                    if await not_now.count() > 0:
                        await not_now.first.click()
                        await page.wait_for_timeout(2000)
                except:
                    pass

            except Exception:
                print('⚠️ 자동 로그인 실패. 이미 로그인되어 있을 수 있습니다.')

            # 게시글 로드
            print('\n게시글 로딩...')
            await page.goto(post_url, wait_until='domcontentloaded', timeout=60000)
            await page.wait_for_timeout(5000)

            print('\n좋아요 버튼 찾는 중...')

            # 좋아요 버튼 찾기 및 클릭
            found_selector = None
            like_selectors = [
                'a[href*="liked_by"]',
                'a:has-text("likes")',
                'a:has-text("좋아요")',
            ]

            for selector in like_selectors:
                count = await page.locator(selector).count()
                if count > 0:
                    print(f'✓ 발견: {selector}')
                    found_selector = selector
                    break

            if found_selector:
                try:
                    await page.locator(found_selector).first.click(timeout=5000)
                    print('✓ 좋아요 모달 열림!')
                    await page.wait_for_timeout(3000)
                except Exception as e:
                    print(f'⚠️ 자동 클릭 실패: {e}')

            # 사용자 목록 스크롤 및 수집
            print('\n사용자 목록 수집 중...')
            liked_users = set()
            previous_count = 0
            no_change_count = 0
            max_scrolls = 200

            for scroll_num in range(max_scrolls):
                # 모달 내 사용자 링크 수집
                user_links = await page.locator('a[href^="/"][href$="/"]').all()

                for link in user_links:
                    try:
                        href = await link.get_attribute('href')
                        if href and href != '/':
                            username = href.strip('/').split('/')[0]
                            if (username and
                                not username.startswith('p/') and
                                not username.startswith('explore') and
                                not username.startswith('direct') and
                                not username.startswith('legal')):
                                liked_users.add(username.lower())
                    except:
                        pass

                # 진행상황 출력
                if scroll_num % 10 == 0:
                    print(f'스크롤 {scroll_num + 1}/{max_scrolls} - 수집: {len(liked_users)}명')

                # 변화 없으면 종료
                if len(liked_users) == previous_count:
                    no_change_count += 1
                    if no_change_count >= 5:
                        print('\n더 이상 새로운 사용자가 없습니다.')
                        break
                else:
                    no_change_count = 0

                previous_count = len(liked_users)

                # 스크롤
                try:
                    await page.keyboard.press('End')
                    await page.wait_for_timeout(800)
                except:
                    pass

            print(f'\n✓ 총 {len(liked_users)}명의 사용자 수집 완료!')
            return liked_users

        except Exception as e:
            print(f'\n❌ 오류 발생: {e}')
            import traceback
            traceback.print_exc()
            return set()
        finally:
            print('\n브라우저를 5초 후 종료합니다...')
            await page.wait_for_timeout(5000)
            await browser.close()


def update_excel_file(liked_users, excel_path, password, output_path):
    """엑셀 파일의 미션 이행 여부를 업데이트합니다."""
    print(f'\n{"="*60}')
    print('엑셀 파일 업데이트 중...')
    print(f'{"="*60}')

    try:
        # 암호화된 파일 열기
        print('\n파일 복호화 중...')
        encrypted_file = open(excel_path, 'rb')
        decrypted = io.BytesIO()
        office_file = msoffcrypto.OfficeFile(encrypted_file)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
        encrypted_file.close()

        # 복호화된 파일 읽기
        decrypted.seek(0)
        wb = openpyxl.load_workbook(decrypted)
        ws = wb.active

        print(f'시트 이름: {ws.title}')
        print(f'총 행 수: {ws.max_row}')

        # 헤더 찾기 (3번째 행에 헤더가 있음)
        headers = [cell.value for cell in ws[3]]
        print(f'\n헤더: {headers}')

        # ID 컬럼과 미션 이행 여부 컬럼 찾기
        id_col = None
        mission_col = None

        for idx, header in enumerate(headers, start=1):
            if header == 'ID':
                id_col = idx
            elif header == '미션 이행 여부':
                mission_col = idx

        if not id_col or not mission_col:
            print(f'❌ 필요한 컬럼을 찾을 수 없습니다. ID: {id_col}, 미션: {mission_col}')
            return

        print(f'\nID 컬럼: {id_col}번째 (열 {openpyxl.utils.get_column_letter(id_col)})')
        print(f'미션 이행 여부 컬럼: {mission_col}번째 (열 {openpyxl.utils.get_column_letter(mission_col)})')

        # 데이터 업데이트 (4번째 행부터 시작)
        updated_count = 0
        matched_count = 0

        print('\n미션 이행 여부 업데이트 중...')

        for row_idx in range(4, ws.max_row + 1):
            instagram_id = ws.cell(row=row_idx, column=id_col).value

            if instagram_id:
                # Instagram ID 정제 (@ 제거, 소문자 변환, 공백 제거)
                clean_id = str(instagram_id).strip().lower().replace('@', '')

                # 좋아요 목록에 있는지 확인
                if clean_id in liked_users:
                    ws.cell(row=row_idx, column=mission_col).value = 'O'
                    matched_count += 1
                else:
                    ws.cell(row=row_idx, column=mission_col).value = 'X'

                updated_count += 1

        print(f'\n업데이트 완료:')
        print(f'  - 총 업데이트된 행: {updated_count}')
        print(f'  - 좋아요 매칭: {matched_count}')
        print(f'  - 매칭 안됨: {updated_count - matched_count}')

        # 파일 저장
        print(f'\n파일 저장 중: {output_path}')
        wb.save(output_path)
        wb.close()

        print(f'\n✓ 파일이 저장되었습니다: {output_path}')
        print(f'{"="*60}')

    except Exception as e:
        print(f'\n❌ 엑셀 업데이트 오류: {e}')
        import traceback
        traceback.print_exc()


async def main():
    """메인 함수"""
    print('\n자일리톨 이벤트 좋아요 확인 프로그램')

    # 1. Instagram 좋아요 목록 수집
    liked_users = await get_instagram_likes(INSTAGRAM_POST_URL)

    if not liked_users:
        print('\n❌ 좋아요 사용자를 수집하지 못했습니다.')
        return

    print(f'\n수집된 사용자 샘플 (처음 10명):')
    for idx, user in enumerate(list(liked_users)[:10], 1):
        print(f'  {idx}. {user}')

    # 2. 엑셀 파일 업데이트
    update_excel_file(liked_users, EXCEL_FILE_PATH, EXCEL_PASSWORD, OUTPUT_FILE_PATH)

    print('\n모든 작업이 완료되었습니다!')


if __name__ == '__main__':
    asyncio.run(main())
