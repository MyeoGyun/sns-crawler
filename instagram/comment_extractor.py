"""
Instagram 댓글 수집기
사용법: python instagram/comment_extractor.py
"""

import asyncio
import os
import sys
import time
import csv
import random
import re
from datetime import datetime
from playwright.async_api import async_playwright
import threading

# 설정
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_FILE = os.path.join(BASE_DIR, ".env")

# DOM 셀렉터 후보
COMMENT_ITEM_SELECTORS = [
    # 게시글 본문
    "article ul li",
    "article div ul li",
    # 피드에서 모달로 열렸을 때
    "div[role='dialog'] article ul li",
    "div[role='dialog'] ul li",
    # 혹시 모를 기타 케이스
    "section ul li",
    "main ul li",
]

LOAD_MORE_BUTTON_SELECTORS = [
    # 전체 댓글 더보기
    'button:has-text("댓글 더 보기")',
    'button:has-text("View all comments")',
    'button:has-text("View more comments")',
    'button:has-text("모두 보기")',
    'button:has-text("Load more comments")',
    'div[role="button"]:has-text("댓글 더 보기")',
    'span:has-text("댓글 더 보기")',
    # 숨김 댓글 / 추가 댓글
    'button:has-text("댓글")',
]

REPLY_EXPANDER_SELECTORS = [
    'button:has-text("답글 보기")',
    'button:has-text("답글")',
    'button:has-text("View replies")',
    'button:has-text("Hide replies")',
    'button:has-text("View all replies")',
]

# -----------------------
# 필수 라이브러리 확인
# -----------------------
def check_dependencies():
    """필수 라이브러리 설치 여부 확인"""
    print(f"\n{'═' * 70}")
    print(f"  📦 필수 라이브러리 확인 중...")
    print(f"{'═' * 70}\n")

    required = {
        'playwright': 'playwright',
        'openpyxl': 'openpyxl',
        'Pillow': 'PIL',
    }

    missing = []
    installed = []

    for display_name, import_name in required.items():
        try:
            module = __import__(import_name)
            version = getattr(module, '__version__', '알 수 없음')
            installed.append(f"  ✅ {display_name:20s} {version}")
        except ImportError:
            missing.append(display_name)

    for msg in installed:
        print(msg)

    if missing:
        print(f"\n{'─' * 70}")
        print(f"  ❌ 다음 라이브러리가 설치되지 않았습니다:")
        for lib in missing:
            print(f"      • {lib}")
        print(f"\n  설치 명령어:")
        print(f"      pip install {' '.join(missing)}")
        print(f"{'─' * 70}\n")
        sys.exit(1)

    print(f"\n{'─' * 70}")
    print(f"  ✅ 모든 필수 라이브러리가 설치되어 있습니다")
    print(f"{'─' * 70}\n")


# -----------------------
# 쿠키 관리
# -----------------------
def get_cookies_from_user():
    """사용자로부터 쿠키 값을 직접 입력받기"""
    print(f"\n{'═' * 70}")
    print(f"  🔑 Instagram 쿠키 입력")
    print(f"{'═' * 70}")
    print(f"\n  📋 Instagram 쿠키 가져오기:")
    print(f"      1. Chrome에서 instagram.com 로그인")
    print(f"      2. F12 > Application > Cookies > https://www.instagram.com")
    print(f"      3. 다음 쿠키 값을 복사:")
    print(f"         • sessionid (필수)")
    print(f"         • csrftoken (필수)")
    print(f"         • ds_user_id (선택)")
    print(f"\n{'─' * 70}\n")

    sessionid = input("  sessionid: ").strip()
    if not sessionid:
        return None

    csrftoken = input("  csrftoken: ").strip()
    if not csrftoken:
        return None

    ds_user_id = input("  ds_user_id (엔터 = 건너뛰기): ").strip()

    cookies = {
        "SESSIONID": sessionid,
        "CSRFTOKEN": csrftoken,
    }

    if ds_user_id:
        cookies["DS_USER_ID"] = ds_user_id

    return cookies


def load_cookies_from_env():
    """Load cookies from .env file"""
    if not os.path.exists(ENV_FILE):
        print(f"  ❌ .env 파일을 찾을 수 없습니다: {ENV_FILE}")
        return None

    cookies = {}
    with open(ENV_FILE, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if '=' in line:
                key, value = line.split('=', 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")

                if key.startswith('INSTAGRAM_'):
                    cookie_name = key.replace('INSTAGRAM_', '')
                    cookies[cookie_name] = value

    if not cookies.get('SESSIONID') or not cookies.get('CSRFTOKEN'):
        print(f"  ❌ .env 파일에 필수 쿠키가 없습니다 (INSTAGRAM_SESSIONID, INSTAGRAM_CSRFTOKEN)")
        return None

    print(f"  ✅ .env 파일에서 {len(cookies)}개 쿠키 로드 완료")
    return cookies


def save_to_env_file(cookies):
    """Save cookies to .env file"""
    print(f"\n{'─' * 70}")
    print(f"  💾 쿠키를 .env 파일로 저장 중...")

    env_content = "# Instagram 인증 쿠키\n"
    for key, value in cookies.items():
        env_content += f"INSTAGRAM_{key}={value}\n"

    with open(ENV_FILE, 'w', encoding='utf-8') as f:
        f.write(env_content)

    print(f"  ✅ .env 파일 저장 완료: {ENV_FILE}")
    print(f"{'─' * 70}\n")


async def load_cookies_to_context(context, cookies):
    """Playwright context에 쿠키 로드"""
    try:
        print(f"\n{'─' * 70}")
        print(f"  🔑 쿠키 로드 중...")

        cookie_list = []

        if cookies.get("SESSIONID"):
            cookie_list.append({
                "name": "sessionid",
                "value": cookies["SESSIONID"],
                "domain": ".instagram.com",
                "path": "/",
            })

        if cookies.get("CSRFTOKEN"):
            cookie_list.append({
                "name": "csrftoken",
                "value": cookies["CSRFTOKEN"],
                "domain": ".instagram.com",
                "path": "/",
            })

        if cookies.get("DS_USER_ID"):
            cookie_list.append({
                "name": "ds_user_id",
                "value": cookies["DS_USER_ID"],
                "domain": ".instagram.com",
                "path": "/",
            })

        await context.add_cookies(cookie_list)
        print(f"  ✅ {len(cookie_list)}개 쿠키 로드 완료")
        print(f"{'─' * 70}\n")
        return True

    except Exception as e:
        print(f"  ❌ 쿠키 로드 실패: {e}")
        print(f"{'─' * 70}\n")
        return False


# -----------------------
# DOM 구조 분석 (디버깅용)
# -----------------------
async def analyze_dom_structure(page):
    """실제 Instagram DOM 구조를 분석하여 출력"""
    print(f"\n{'═' * 70}")
    print(f"  🔍 DOM 구조 분석 중...")
    print(f"{'═' * 70}\n")

    # 댓글 컨테이너 찾기
    selectors_to_test = [
        'ul[role="list"]',  # 댓글 리스트
        'div[role="dialog"]',  # 모달
        'article',  # 게시글
        'ul',  # 일반 리스트
        'li',  # 리스트 아이템
    ]

    for selector in selectors_to_test:
        count = await page.locator(selector).count()
        print(f"  {selector:30s} : {count}개 발견")

    # 첫 번째 댓글 구조 분석
    print(f"\n{'─' * 70}")
    print(f"  📝 댓글 요소 상세 분석:")
    print(f"{'─' * 70}\n")

    # 다양한 패턴으로 댓글 찾기 시도
    comment_patterns = [
        'ul li',
        'div[role="button"]',
        'span:has-text("좋아요")',
        'time[datetime]',
        'h2 + div + div ul li',  # 댓글 리스트 아이템
        'span[dir="auto"]',  # 텍스트 span
    ]

    for pattern in comment_patterns:
        count = await page.locator(pattern).count()
        if count > 0:
            print(f"  ✓ {pattern:40s} : {count}개")
            if count <= 5:  # 5개 이하면 텍스트도 출력
                for i in range(min(count, 3)):
                    try:
                        text = await page.locator(pattern).nth(i).inner_text()
                        print(f"      [{i}] {text[:100]}")
                    except:
                        pass

    # 페이지 HTML 일부 저장 (디버깅용)
    try:
        html_content = await page.content()
        debug_file = "instagram_debug.html"
        with open(debug_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        print(f"\n  💾 전체 HTML 저장: {debug_file}")
    except:
        pass

    print(f"\n{'═' * 70}\n")


# -----------------------
# 헬퍼 유틸
# -----------------------
def normalize_text(value):
    """불필요한 공백 제거"""
    if not value:
        return ""
    return re.sub(r"\s+", " ", value).strip()


def build_comment_key(comment):
    """중복 방지를 위한 고유 키 생성"""
    username = normalize_text(comment.get("username", "")).lower()
    text = normalize_text(comment.get("text", ""))
    timestamp = comment.get("timestamp", "")
    return f"{username}|{text}|{timestamp}"


def convert_count_to_number(count_str):
    """
    Instagram 숫자 표기(K, M, B)를 실제 숫자로 변환

    Args:
        count_str: '18K', '1.2M', '500', 'N/A' 등의 문자열

    Returns:
        int/float: 변환된 숫자, 또는 원본 문자열 (변환 불가 시)

    Examples:
        '18K' -> 18000
        '1.2M' -> 1200000
        '500' -> 500
        'N/A' -> 'N/A'
        '0' -> 0
    """
    if not count_str or count_str in ['N/A', 'n/a']:
        return 'N/A'

    count_str = str(count_str).strip().upper()

    # 이미 순수 숫자인 경우
    try:
        return int(count_str.replace(',', ''))
    except ValueError:
        pass

    # K, M, B 변환
    multipliers = {
        'K': 1_000,
        'M': 1_000_000,
        'B': 1_000_000_000,
    }

    for suffix, multiplier in multipliers.items():
        if suffix in count_str:
            try:
                # '18K', '1.2M' 등에서 숫자 부분 추출
                number_part = count_str.replace(suffix, '').replace(',', '').strip()
                number = float(number_part)
                result = number * multiplier
                # 정수로 변환 가능하면 정수로
                return int(result) if result == int(result) else result
            except ValueError:
                pass

    # 변환 실패 시 원본 반환
    return count_str


async def click_first_available(page, selectors):
    """셀렉터 목록 중 클릭 가능한 첫 요소 클릭"""
    for selector in selectors:
        try:
            locator = page.locator(selector)
            count = await locator.count()
            if count == 0:
                continue
            target = locator.nth(0)
            try:
                await target.scroll_into_view_if_needed()
            except:
                pass
            await target.click(timeout=3000)
            return True
        except Exception:
            continue
    return False


async def click_load_more_comments(page):
    """댓글/답글 더보기 버튼 클릭"""
    clicked = await click_first_available(page, LOAD_MORE_BUTTON_SELECTORS)
    reply_clicked = await click_first_available(page, REPLY_EXPANDER_SELECTORS)
    return clicked or reply_clicked


async def wait_for_comment_section(page, timeout=20000):
    """댓글 섹션이 렌더링될 때까지 대기"""
    for selector in COMMENT_ITEM_SELECTORS:
        try:
            await page.wait_for_selector(selector, timeout=timeout)
            return True
        except Exception:
            continue
    return False


async def collect_visible_comments(page, seen_ids):
    """
    time[datetime]을 기준으로 컨테이너를 찾고,
    본문은 span[dir="auto"] 내부 전체 텍스트를 추출한다.
    """

    raw_comments = await page.evaluate(
        """
        () => {
            const relativeTimePattern = '\\\\d+\\\\s*(분|시간|일|주|개월|년)';
            const relativeTimeRegex = new RegExp(`^${relativeTimePattern}$`);
            const relativeTimeInline = new RegExp(relativeTimePattern, 'g');
            const results = [];

    const commentLinks = document.querySelectorAll(
        "a[href*='/p/'][href*='/c/']"
    );
    const processedRoots = new WeakSet();

    commentLinks.forEach((link) => {
        const timeEl = link.querySelector('time[datetime]');
        if (!timeEl) {
            return;
        }
        const timestamp = timeEl.getAttribute('datetime') || '';
        if (!timestamp) {
            return;
        }

        // DEBUG: 첫 3개 timestamp 출력
        if (results.length < 3) {
            console.log('[DEBUG] timestamp:', timestamp, 'type:', typeof timestamp);
        }

        let container = link.parentElement;
        let chosen = null;
        for (let i = 0; i < 25 && container; i += 1) {
            const hasUser = container.querySelector('a._a6hd span, a._a6hd');
            const hasBody = container.querySelector('span[dir="auto"]');
            const hasTime = container.querySelector('time[datetime]');
            const linkCount = container.querySelectorAll(
                "a[href*='/p/'][href*='/c/']"
            ).length;
            if (hasUser && hasBody && hasTime && linkCount === 1) {
                chosen = container;
            }
            if (linkCount > 1) {
                break;
            }
            container = container.parentElement;
        }

        container = chosen || container;

        if (!container || processedRoots.has(container)) {
            return;
        }
        processedRoots.add(container);

        // 사용자명
        const usernameNode =
            container.querySelector('a._a6hd span') ||
            container.querySelector('a._a6hd') ||
            container.querySelector('a[role="link"][tabindex="0"] span') ||
            container.querySelector('a[role="link"][tabindex="0"]');
        const username = usernameNode ? usernameNode.textContent.trim() : '';

                // 댓글 본문
                const metaPatterns = [
                    /^수정됨$/,
                    /^답글.*모두 보기$/,
                    /^좋아요\s*\d*/,
                    /^좋아요$/,
                    /^답글 달기$/,
                ];

                const textPieces = [];
                container.querySelectorAll('span[dir="auto"]').forEach((span) => {
                    if (span.closest('button')) {
                        return;
                    }
                    const raw = (span.textContent || '').trim();
                    if (!raw) {
                        return;
                    }
                    if (username && raw === username) {
                        return;
                    }
                    if (relativeTimeRegex.test(raw)) {
                        return;
                    }
                    if (metaPatterns.some((regex) => regex.test(raw))) {
                        return;
                    }
                    if (/^(좋아요|답글|Reply|View|좋아요를 누른 사람이)/.test(raw)) {
                        return;
                    }
                    textPieces.push(raw);
                });
                let text = textPieces.join(' ').replace(/\\s+/g, ' ').trim();

                if (!text) {
                    let fallback = (container.innerText || '').trim();
                    if (username && fallback.startsWith(username)) {
                        fallback = fallback.slice(username.length).trim();
                    }
                    fallback = fallback
                        .replace(relativeTimeInline, '')
                        .replace(/좋아요\\s*\\d*/gi, '')
                        .replace(/답글( 달기)?/gi, '')
                        .replace(/•/g, '')
                        .replace(/\\s+/g, ' ')
                        .trim();
                    if (!metaPatterns.some((regex) => regex.test(fallback))) {
                        text = fallback;
                    }
                }

                // 좋아요/답글 수
                let likes = '';
                let replies = '';
                container.querySelectorAll('button').forEach((btn) => {
                    const label = (btn.textContent || '').trim();
                    if (!label) {
                        return;
                    }
                    const numbers = label.match(/\\d+/);
                    if (/좋아요|likes?/i.test(label) && numbers && !likes) {
                        likes = numbers[0];
                    } else if (/답글|repl/i.test(label) && numbers && !replies) {
                        replies = numbers[0];
                    }
                });

                if (!username && !text) {
                    return;
                }

                results.push({
                    username,
                    text,
                    timestamp,
                    likes,
                    replies,
                });
            });

            return results;
        }
        """
    )

    harvested = []
    for comment in raw_comments or []:
        comment["username"] = normalize_text(comment.get("username", ""))
        comment["text"] = normalize_text(comment.get("text", ""))
        comment["timestamp"] = comment.get("timestamp", "")
        comment["likes"] = (comment.get("likes") or "0").strip() or "0"
        comment["replies"] = (comment.get("replies") or "0").strip() or "0"

        # 프로필 링크 생성
        comment["profile_link"] = f"https://www.instagram.com/{comment['username']}/" if comment["username"] else ""

        comment_id = build_comment_key(comment)
        if not comment_id or comment_id in seen_ids:
            continue

        seen_ids.add(comment_id)
        harvested.append(comment)

    return harvested


# -----------------------
# 프로필 메트릭 수집
# -----------------------
async def collect_user_profile_metrics(page, username, retry_count=3):
    """
    사용자 프로필 메트릭 수집: 게시물수, 팔로워수, 팔로잉수, 비공개계정여부

    Args:
        page: Playwright page object
        username: Instagram username
        retry_count: Number of retries on failure

    Returns:
        dict: {
            'posts_count': str,
            'followers_count': str,
            'following_count': str,
            'is_private': str ('O' or 'X'),
            'success': bool,
            'error': str
        }
    """
    result = {
        'posts_count': '0',
        'followers_count': '0',
        'following_count': '0',
        'is_private': 'X',
        'success': False,
        'error': ''
    }

    if not username:
        result['error'] = 'Empty username'
        return result

    profile_url = f"https://www.instagram.com/{username}/"

    for attempt in range(retry_count):
        try:
            # Navigate to profile
            await page.goto(profile_url, wait_until='domcontentloaded', timeout=30000)
            await page.wait_for_timeout(random.uniform(2.0, 4.0) * 1000)

            # Method 1: Try meta description (most reliable)
            try:
                meta_desc = await page.locator('meta[name="description"]').get_attribute('content')
                if meta_desc:
                    import re

                    # DEBUG: Print meta description
                    print(f"\n      [DEBUG] Meta description: {meta_desc[:200]}...")

                    # Extract numbers before keywords (English)
                    followers_match = re.search(r'([\d,\.KMB]+)\s*Followers', meta_desc, re.IGNORECASE)
                    following_match = re.search(r'([\d,\.KMB]+)\s*Following', meta_desc, re.IGNORECASE)
                    posts_match = re.search(r'([\d,\.KMB]+)\s*Posts', meta_desc, re.IGNORECASE)

                    # Also check Korean
                    if not followers_match:
                        followers_match = re.search(r'팔로워\s*([\d,\.KMB만천백십억]+)명?', meta_desc)
                    if not following_match:
                        following_match = re.search(r'팔로잉?\s*([\d,\.KMB만천백십억]+)명?', meta_desc)
                    if not posts_match:
                        posts_match = re.search(r'게시물\s*([\d,\.KMB만천백십억]+)개?', meta_desc)

                    # DEBUG: Print matches
                    print(f"      [DEBUG] Posts match: {posts_match.group(1) if posts_match else 'None'}")
                    print(f"      [DEBUG] Followers match: {followers_match.group(1) if followers_match else 'None'}")
                    print(f"      [DEBUG] Following match: {following_match.group(1) if following_match else 'None'}")

                    if followers_match:
                        result['followers_count'] = followers_match.group(1).replace(',', '').replace('.', '')
                    if following_match:
                        result['following_count'] = following_match.group(1).replace(',', '').replace('.', '')
                    if posts_match:
                        result['posts_count'] = posts_match.group(1).replace(',', '').replace('.', '')
            except Exception as e:
                print(f"      [DEBUG] Meta parsing error: {e}")
                pass  # Silently fail and try next method

            # Method 2: Try direct DOM selectors (backup)
            if result['followers_count'] == '0':
                try:
                    # Look for stat containers - common Instagram pattern
                    stats = await page.locator('header section ul li').all()
                    print(f"      [DEBUG] DOM stats found: {len(stats)} elements")

                    if len(stats) >= 3:
                        # Usually: posts, followers, following (in that order)
                        posts_text = await stats[0].locator('span').first.inner_text()
                        followers_text = await stats[1].locator('span').first.inner_text()
                        following_text = await stats[2].locator('span').first.inner_text()

                        # DEBUG: Print raw text
                        print(f"      [DEBUG] DOM stats[0] (posts): '{posts_text}'")
                        print(f"      [DEBUG] DOM stats[1] (followers): '{followers_text}'")
                        print(f"      [DEBUG] DOM stats[2] (following): '{following_text}'")

                        # Clean numbers (remove commas, spaces)
                        import re
                        result['posts_count'] = re.sub(r'[^0-9KMB]', '', posts_text)
                        result['followers_count'] = re.sub(r'[^0-9KMB]', '', followers_text)
                        result['following_count'] = re.sub(r'[^0-9KMB]', '', following_text)
                except Exception as e:
                    print(f"      [DEBUG] DOM parsing error: {e}")
                    pass  # Silently fail

            # Check if private account
            try:
                is_private_en = await page.locator('h2:has-text("This Account is Private")').count()
                is_private_ko = await page.locator('h2:has-text("비공개 계정")').count()
                result['is_private'] = 'O' if (is_private_en > 0 or is_private_ko > 0) else 'X'
            except:
                result['is_private'] = 'X'

            # Mark success if we got at least one metric
            if (result['followers_count'] != '0' or
                result['posts_count'] != '0' or
                result['following_count'] != '0'):
                result['success'] = True
                # DEBUG: Final result
                print(f"      [DEBUG] Final result - Posts: {result['posts_count']}, Followers: {result['followers_count']}, Following: {result['following_count']}")
                return result

            # If failed, wait before retry
            if attempt < retry_count - 1:
                await page.wait_for_timeout(random.uniform(3.0, 5.0) * 1000)

        except Exception as e:
            result['error'] = str(e)
            if attempt < retry_count - 1:
                await page.wait_for_timeout(random.uniform(5.0, 8.0) * 1000)

    return result


async def enrich_comments_with_profile_data(page, comments, max_profiles=None):
    """
    댓글 데이터에 사용자 프로필 메트릭 추가

    Args:
        page: Playwright page
        comments: List of comment dicts
        max_profiles: Maximum number of profiles to scrape (None = all)

    Returns:
        List of enriched comments
    """
    print(f"\n{'═' * 70}")
    print(f"  👤 사용자 프로필 정보 수집 시작")
    print(f"{'═' * 70}")

    # Get unique usernames
    unique_users = {}
    for comment in comments:
        username = comment.get('username', '').strip()
        if username and username not in unique_users:
            unique_users[username] = None

    total_users = len(unique_users)
    if max_profiles:
        total_users = min(total_users, max_profiles)

    print(f"\n  📊 총 {total_users}명의 프로필 정보 수집 예정")
    print(f"  ⏱️  예상 소요 시간: {total_users * 4 / 60:.1f}분 (프로필당 ~4초)")
    print(f"\n  {'─' * 66}\n")

    processed = 0
    failed = 0

    usernames_to_process = list(unique_users.keys())[:max_profiles] if max_profiles else list(unique_users.keys())

    for username in usernames_to_process:
        processed += 1

        print(f"  [{processed}/{total_users}] @{username} 수집 중...", end=" ", flush=True)

        metrics = await collect_user_profile_metrics(page, username)

        if metrics['success']:
            unique_users[username] = metrics
            print(f"✓ 팔로워: {metrics['followers_count']:>6s} | "
                  f"팔로잉: {metrics['following_count']:>6s} | "
                  f"게시물: {metrics['posts_count']:>6s} | "
                  f"비공개: {metrics['is_private']}")
        else:
            failed += 1
            unique_users[username] = metrics  # Keep failed result
            print(f"✗ 실패: {metrics.get('error', 'Unknown error')[:40]}")

        # Rate limiting: wait between requests
        if processed < total_users:
            wait_time = random.uniform(2.5, 4.5)
            await page.wait_for_timeout(wait_time * 1000)

    print(f"\n  {'─' * 66}")
    print(f"  ✅ 프로필 수집 완료: 성공 {processed - failed}/{total_users}, 실패 {failed}")
    print(f"  {'─' * 66}\n")

    # Merge metrics into comments
    for comment in comments:
        username = comment.get('username', '').strip()
        if username in unique_users and unique_users[username]:
            metrics = unique_users[username]
            if metrics['success']:
                comment['posts_count'] = metrics['posts_count']
                comment['followers_count'] = metrics['followers_count']
                comment['following_count'] = metrics['following_count']
                comment['is_private'] = metrics['is_private']
            else:
                comment['posts_count'] = 'N/A'
                comment['followers_count'] = 'N/A'
                comment['following_count'] = 'N/A'
                comment['is_private'] = 'N/A'
        else:
            comment['posts_count'] = 'N/A'
            comment['followers_count'] = 'N/A'
            comment['following_count'] = 'N/A'
            comment['is_private'] = 'N/A'

    return comments


# -----------------------
# 댓글 파싱
# -----------------------
async def parse_comment(li_element):
    """개별 댓글 요소에서 데이터 추출"""
    try:
        comment_data = {
            "username": "",
            "text": "",
            "timestamp": "",
            "likes": "0",
            "replies": "0",
        }

        # 사용자명 (링크 텍스트)
        try:
            username_link = li_element.locator('a[href^="/"]').first
            comment_data["username"] = await username_link.inner_text()
        except:
            pass

        # 댓글 내용 (span 태그)
        try:
            # 여러 패턴 시도
            text_patterns = [
                'span[dir="auto"]',
                'span',
                'div span',
            ]
            for pattern in text_patterns:
                text_elements = li_element.locator(pattern)
                count = await text_elements.count()
                if count > 0:
                    # 사용자명이 아닌 span 찾기
                    for i in range(count):
                        text = await text_elements.nth(i).inner_text()
                        if text and text != comment_data["username"] and len(text) > 2:
                            comment_data["text"] = text
                            break
                if comment_data["text"]:
                    break
        except:
            pass

        # 타임스탬프
        try:
            time_element = li_element.locator('time[datetime]').first
            comment_data["timestamp"] = await time_element.get_attribute('datetime')
        except:
            # datetime 속성이 없으면 텍스트 사용
            try:
                time_element = li_element.locator('time').first
                comment_data["timestamp"] = await time_element.inner_text()
            except:
                pass

        # 좋아요 수 (버튼 텍스트에서 추출)
        try:
            like_button = li_element.locator('button:has-text("좋아요"), button:has-text("likes")').first
            like_text = await like_button.inner_text()
            numbers = re.findall(r'\d+', like_text)
            if numbers:
                comment_data["likes"] = numbers[0]
        except:
            pass

        # 대댓글 수 (답글 보기 버튼)
        try:
            reply_button = li_element.locator('button:has-text("답글"), button:has-text("repl")').first
            reply_text = await reply_button.inner_text()
            numbers = re.findall(r'\d+', reply_text)
            if numbers:
                comment_data["replies"] = numbers[0]
        except:
            pass

        # 최소한 사용자명과 텍스트가 있어야 유효한 댓글
        if comment_data["username"] or comment_data["text"]:
            return comment_data
        return None

    except Exception as e:
        return None


# -----------------------
# 댓글 추출 (자동 스크롤)
# -----------------------
async def extract_comments_auto(page, post_url, max_scrolls=100):
    """자동 스크롤 모드로 댓글 수집"""
    print(f"\n{'═' * 70}")
    print(f"  🔄 댓글 자동 수집 시작")
    print(f"{'═' * 70}")
    print(f"  📄 게시글: {post_url}")
    print(f"  🔢 최대 스크롤: {max_scrolls}회")
    print(f"{'═' * 70}\n")

    # 게시글 로드
    await page.goto(post_url, wait_until='domcontentloaded', timeout=60000)
    await page.wait_for_timeout(3000)

    # DOM 구조 분석 (디버깅)
    await analyze_dom_structure(page)

    if not await wait_for_comment_section(page):
        print("  ⚠️  댓글 섹션을 찾지 못했습니다. 로그인 여부를 확인해주세요.")
        return []

    comments = []
    seen_ids = set()
    scrolls = 0
    no_change_count = 0

    print(f"  {'─' * 66}")
    print(f"  🔄 스크롤 및 수집 시작")
    print(f"  {'─' * 66}\n")

    for scroll_num in range(1, max_scrolls + 1):
        scrolls = scroll_num

        newly_found = await collect_visible_comments(page, seen_ids)
        if newly_found:
            comments.extend(newly_found)
            no_change_count = 0
            for comment in newly_found[:3]:
                preview = (comment["text"][:60] + "...") if len(comment["text"]) > 60 else comment["text"]
                print(f"  💬 @{comment['username'] or 'unknown'} | {preview}")
        else:
            no_change_count += 1

        print(f"\n  📊 스크롤 #{scroll_num:02d} / {max_scrolls}  |  누적 {len(comments)}개 수집  |  최근 변화 {'O' if newly_found else 'X'}")

        # 더보기/답글 펼치기
        clicked = await click_load_more_comments(page)

        if not clicked:
            # 자연스러운 스크롤
            await page.mouse.wheel(0, random.randint(600, 1200))
            await page.wait_for_timeout(random.uniform(0.8, 1.4) * 1000)
        else:
            await page.wait_for_timeout(random.uniform(1.0, 1.6) * 1000)

        if no_change_count >= 5:
            print(f"\n  🏁 더 이상 새로운 데이터가 감지되지 않아 자동 종료합니다")
            break

    # 마지막으로 한 번 더 수집
    final_batch = await collect_visible_comments(page, seen_ids)
    if final_batch:
        comments.extend(final_batch)

    print(f"\n{'═' * 70}")
    print(f"  ✅ 수집 완료!")
    print(f"  {'─' * 66}")
    print(f"      📝 총 수집 댓글: {len(comments)}개")
    print(f"      🔄 총 스크롤: {scrolls}회")
    print(f"{'═' * 70}\n")

    return comments


# -----------------------
# 댓글 추출 (수동 스크롤)
# -----------------------
async def extract_comments_manual(page, post_url):
    """수동 스크롤 모드 - 사용자가 직접 스크롤"""
    print(f"\n{'═' * 70}")
    print(f"  👆 댓글 수동 수집 모드")
    print(f"{'═' * 70}")
    print(f"  📄 게시글: {post_url}")
    print(f"  💡 브라우저를 직접 스크롤하면 실시간으로 댓글을 수집합니다")
    print(f"  ⏹️  엔터 키를 누르면 수집을 종료합니다")
    print(f"{'═' * 70}\n")

    # 게시글 로드
    await page.goto(post_url, wait_until='domcontentloaded', timeout=60000)
    await page.wait_for_timeout(3000)

    # DOM 구조 분석
    await analyze_dom_structure(page)
    if not await wait_for_comment_section(page):
        print("  ⚠️  댓글 섹션을 찾지 못했습니다. 수동 모드로 계속 대기합니다.")
    else:
        # 기본 댓글/답글 더보기 버튼 한 번 클릭 시도
        await click_load_more_comments(page)

    # 엔터 키 대기 스레드
    stop_flag = [False]

    def wait_for_enter():
        print("  ⏸️  엔터 키를 누르면 수집을 종료합니다...\n")
        input()
        stop_flag[0] = True
        print("\n  🛑 수집 종료 요청됨")

    input_thread = threading.Thread(target=wait_for_enter, daemon=True)
    input_thread.start()

    comments = []
    seen_ids = set()

    print(f"  {'─' * 66}")
    print(f"  👀 실시간 수집 중... (0.5초마다 폴링)")
    print(f"  {'─' * 66}\n")

    while not stop_flag[0]:
        await asyncio.sleep(0.5)

        new_comments = await collect_visible_comments(page, seen_ids)
        if new_comments:
            comments.extend(new_comments)
            print(f"  ➕ 신규 댓글 {len(new_comments)}개 감지 (누적 {len(comments)}개)")
            for comment in new_comments[:3]:
                preview = (comment["text"][:60] + "...") if len(comment["text"]) > 60 else comment["text"]
                print(f"     · @{comment['username'] or 'unknown'} | {preview}")

        # 브라우저가 닫혔는지 체크
        try:
            await page.evaluate("1")
        except:
            print("\n  ⚠️  브라우저가 닫혔습니다")
            break

    print(f"\n{'═' * 70}")
    print(f"  ✅ 수집 완료!")
    print(f"  {'─' * 66}")
    print(f"      📝 총 수집 댓글: {len(comments)}개")
    print(f"{'═' * 70}\n")

    return comments


def collect_follow_target_labels(comments):
    """수집된 댓글에서 팔로우 체크 대상 라벨 목록 추출 (중복 제거, 순서 유지)"""
    labels = []
    for comment in comments:
        follow_map = comment.get("follow_status") or {}
        for label in follow_map.keys():
            if label not in labels:
                labels.append(label)
    return labels


# -----------------------
# 엑셀 저장
# -----------------------
def save_to_excel(comments, output_path):
    """엑셀 파일 저장"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("  ⚠️  openpyxl 라이브러리가 필요합니다: pip install openpyxl Pillow")
        return

    if not comments:
        print("  ⚠️  저장할 데이터 없음")
        return

    print(f"\n{'─' * 70}")
    print(f"  📊 엑셀 파일 생성 중...")
    print(f"{'─' * 70}")

    # 작성일 기준 정렬 (가장 과거의 글이 맨 위로) - 복사본 생성하여 원본 보존
    sorted_comments = sorted(comments, key=lambda x: x.get('timestamp', ''), reverse=False)
    follow_targets = collect_follow_target_labels(sorted_comments)

    wb = Workbook()
    ws = wb.active
    ws.title = "댓글 목록"

    # 헤더
    headers = ["작성자", "프로필 링크", "댓글 내용", "작성 시간", "좋아요 수", "대댓글 수",
               "게시물수", "팔로워수", "팔로잉수", "비공개계정여부"]
    headers.extend([f"팔로우여부({label})" for label in follow_targets])
    ws.append(headers)

    # 헤더 스타일
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 열 너비
    ws.column_dimensions['A'].width = 20  # 작성자
    ws.column_dimensions['B'].width = 50  # 프로필 링크
    ws.column_dimensions['C'].width = 60  # 댓글 내용
    ws.column_dimensions['D'].width = 20  # 작성 시간
    ws.column_dimensions['E'].width = 12  # 좋아요 수
    ws.column_dimensions['F'].width = 12  # 대댓글 수
    ws.column_dimensions['G'].width = 12  # 게시물수
    ws.column_dimensions['H'].width = 12  # 팔로워수
    ws.column_dimensions['I'].width = 12  # 팔로잉수
    ws.column_dimensions['J'].width = 15  # 비공개계정여부
    if follow_targets:
        from openpyxl.utils import get_column_letter
        for idx, _ in enumerate(follow_targets, start=11):
            ws.column_dimensions[get_column_letter(idx)].width = 18  # 팔로우 여부

    # 데이터 행
    for idx, comment in enumerate(sorted_comments, start=2):
        timestamp = comment.get("timestamp", "")
        formatted = timestamp
        parsed_dt = None
        if timestamp:
            try:
                # 먼저 ISO 형식 시도
                parsed_dt = datetime.fromisoformat(timestamp.replace("Z", "+00:00")).replace(
                    tzinfo=None
                )
                formatted = parsed_dt.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                # ISO 형식 실패 시 Unix timestamp로 시도
                try:
                    # timestamp가 숫자 문자열인지 확인 (예: "45.980")
                    timestamp_float = float(timestamp)
                    # Unix timestamp(밀리초)를 datetime으로 변환
                    if timestamp_float > 1000000000000:  # 밀리초 형식 (13자리)
                        parsed_dt = datetime.fromtimestamp(timestamp_float / 1000)
                    else:  # 초 형식
                        parsed_dt = datetime.fromtimestamp(timestamp_float)
                    formatted = parsed_dt.strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    parsed_dt = None
                    formatted = timestamp

        row_data = [
            comment.get("username", ""),
            comment.get("profile_link", ""),
            comment.get("text", ""),
            formatted,
            comment.get("likes", "0"),
            comment.get("replies", "0"),
            comment.get("posts_count", "N/A"),
            comment.get("followers_count", "N/A"),
            comment.get("following_count", "N/A"),
            comment.get("is_private", "N/A"),
        ]
        follow_map = comment.get("follow_status") or {}
        for label in follow_targets:
            row_data.append(follow_map.get(label, ""))
        ws.append(row_data)

        # 작성 시간 셀에 datetime 객체로 저장하면 Excel에서 형식 지정 가능
        time_cell = ws.cell(row=idx, column=4)
        if parsed_dt:
            time_cell.value = parsed_dt
            time_cell.number_format = "yyyy-mm-dd hh:mm:ss"
        else:
            time_cell.value = formatted

        # 게시물수, 팔로워수, 팔로잉수를 숫자로 변환하여 셀에 적용
        for col_idx, field in [(7, 'posts_count'), (8, 'followers_count'), (9, 'following_count')]:
            cell = ws.cell(row=idx, column=col_idx)
            value = comment.get(field, 'N/A')
            converted = convert_count_to_number(value)

            if isinstance(converted, (int, float)):
                # 숫자로 변환 성공 - Excel 숫자 타입으로 저장
                cell.value = converted
                cell.number_format = '#,##0'  # 천단위 구분 포맷
            else:
                # 'N/A' 등 - 문자열로 유지
                cell.value = converted

        # 텍스트 정렬
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # 저장
    wb.save(output_path)

    print(f"\n{'─' * 70}")
    print(f"  💾 엑셀 파일 저장 완료!")
    print(f"  {'─' * 66}")
    print(f"      📁 파일명: {output_path}")
    print(f"      📊 레코드: {len(comments)}개")
    print(f"{'─' * 70}\n")


# -----------------------
# CSV 저장
# -----------------------
def save_to_csv(comments, output_path):
    """CSV 저장"""
    if not comments:
        print("  ⚠️  저장할 데이터 없음")
        return

    # 작성일 기준 정렬 (가장 과거의 글이 맨 위로)
    sorted_comments = sorted(comments, key=lambda x: x.get('timestamp', ''), reverse=False)
    follow_targets = collect_follow_target_labels(sorted_comments)

    fieldnames = ["작성자", "프로필 링크", "댓글 내용", "작성 시간", "좋아요 수", "대댓글 수",
                  "게시물수", "팔로워수", "팔로잉수", "비공개계정여부"]
    fieldnames.extend([f"팔로우여부({label})" for label in follow_targets])

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for comment in sorted_comments:
            timestamp = comment.get("timestamp", "")
            formatted = timestamp
            if timestamp:
                try:
                    # 먼저 ISO 형식 시도
                    parsed_dt = datetime.fromisoformat(timestamp.replace("Z", "+00:00")).replace(
                        tzinfo=None
                    )
                    formatted = parsed_dt.strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    # ISO 형식 실패 시 Unix timestamp로 시도
                    try:
                        timestamp_float = float(timestamp)
                        if timestamp_float > 1000000000000:  # 밀리초
                            parsed_dt = datetime.fromtimestamp(timestamp_float / 1000)
                        else:  # 초
                            parsed_dt = datetime.fromtimestamp(timestamp_float)
                        formatted = parsed_dt.strftime("%Y-%m-%d %H:%M:%S")
                    except Exception:
                        formatted = timestamp

            # 게시물수, 팔로워수, 팔로잉수를 숫자로 변환
            posts = convert_count_to_number(comment.get("posts_count", "N/A"))
            followers = convert_count_to_number(comment.get("followers_count", "N/A"))
            following = convert_count_to_number(comment.get("following_count", "N/A"))

            mapped_comment = {
                "작성자": comment.get("username", ""),
                "프로필 링크": comment.get("profile_link", ""),
                "댓글 내용": comment.get("text", ""),
                "작성 시간": formatted,
                "좋아요 수": comment.get("likes", "0"),
                "대댓글 수": comment.get("replies", "0"),
                "게시물수": posts,
                "팔로워수": followers,
                "팔로잉수": following,
                "비공개계정여부": comment.get("is_private", "N/A"),
            }
            follow_map = comment.get("follow_status") or {}
            for label in follow_targets:
                mapped_comment[f"팔로우여부({label})"] = follow_map.get(label, "")
            writer.writerow(mapped_comment)

    print(f"\n{'─' * 70}")
    print(f"  💾 CSV 파일 저장 완료!")
    print(f"  {'─' * 66}")
    print(f"      📁 파일명: {output_path}")
    print(f"      📊 레코드: {len(comments)}개")
    print(f"{'─' * 70}\n")


# -----------------------
# 팔로우 여부 확인
# -----------------------
FOLLOWING_LINK_SELECTORS = [
    "header a[href$='/following/']",
    "header section ul li:nth-child(3) a",
    'header a:has-text("팔로우")',
    'header a:has-text("팔로잉")',
    'header a:has-text("following")',
]

FOLLOW_SEARCH_INPUT_SELECTORS = [
    'div[role="dialog"] input[placeholder="검색"]',
    'div[role="dialog"] input[aria-label="검색"]',
    'div[role="dialog"] input[aria-label="Search"]',
]


def normalize_target_username(value):
    """@, URL 등을 제거해 Instagram 사용자명 형태로 정규화"""
    if not value:
        return ""
    normalized = value.strip()
    normalized = normalized.replace("https://www.instagram.com/", "")
    normalized = normalized.replace("http://www.instagram.com/", "")
    normalized = normalized.replace("https://instagram.com/", "")
    normalized = normalized.replace("http://instagram.com/", "")
    normalized = normalized.split("?")[0]
    normalized = normalized.strip("/")
    normalized = normalized.lstrip("@")
    normalized = re.sub(r'\s+', '', normalized)
    return normalized.strip().lower()


def build_profile_url(profile_link, fallback_username=None):
    """프로필 링크가 없을 경우 사용자명으로 URL 구성"""
    base = profile_link or ""
    if not base and fallback_username:
        slug = normalize_target_username(fallback_username)
        if slug:
            base = f"https://www.instagram.com/{slug}/"

    if not base:
        return None

    if base.startswith("//"):
        base = f"https:{base}"
    elif base.startswith("/"):
        base = f"https://www.instagram.com{base}"

    if not base.startswith("http"):
        base = f"https://www.instagram.com/{base.lstrip('/')}"

    base = base.split("?")[0]
    if not base.endswith("/"):
        base += "/"
    return base


def parse_follow_target_input(raw_input):
    """쉼표로 구분된 팔로우 대상 문자열을 spec 리스트로 변환
    형식 예시:
      • roo_lab         → 검색어=라벨=매칭 텍스트 동일
      • roo_lab|이로운 연구소  → 검색어 'roo_lab', 표시 텍스트 '이로운 연구소'
      • roo_lab=>이로운연구소  (구분자 |, =>, :: 지원)
    """
    if not raw_input:
        return []

    def split_chunk(chunk):
        for delimiter in ["|", "=>", "::"]:
            if delimiter in chunk:
                return chunk.split(delimiter, 1)
        return chunk, ""

    specs = []
    seen_labels = set()
    for chunk in raw_input.split(","):
        raw_chunk = chunk.strip()
        if not raw_chunk:
            continue
        search_part, display_part = split_chunk(raw_chunk)
        search_term = search_part.strip()
        display_term = display_part.strip()
        if not search_term:
            continue
        label = display_term or search_term
        label_key = label.lower()
        if label_key in seen_labels:
            continue
        seen_labels.add(label_key)
        lower_search = search_term.lower()
        lower_display = display_term.lower() if display_term else ""
        specs.append({
            "label": label,
            "search_term": search_term,
            "normalized": normalize_target_username(search_term),
            "raw_lower": lower_search,
            "raw_lower_compact": re.sub(r'\s+', '', lower_search),
            "display_lower": lower_display,
            "display_compact": re.sub(r'\s+', '', lower_display) if display_term else "",
        })
    return specs


async def open_following_modal(page):
    """프로필 페이지에서 팔로잉 목록 모달 오픈"""
    for selector in FOLLOWING_LINK_SELECTORS:
        locator = page.locator(selector)
        try:
            if await locator.count() > 0:
                await locator.first.click()
                await page.wait_for_timeout(1500)
                dialog = page.locator('div[role="dialog"]').last
                await dialog.wait_for(timeout=5000)
                return dialog
        except Exception:
            continue
    return None


async def wait_for_follow_search_input(page):
    """팔로우 모달 안의 검색 입력창을 탐색"""
    for selector in FOLLOW_SEARCH_INPUT_SELECTORS:
        locator = page.locator(selector)
        if await locator.count() > 0:
            try:
                await locator.first.wait_for(timeout=4000)
                return locator.first
            except Exception:
                continue
    return None


async def detect_follow_targets_in_dialog(dialog, target_spec):
    """모달 내 anchor/text를 검사하여 대상 계정 존재 여부 확인"""
    handle = await dialog.element_handle()
    if not handle:
        return False

    try:
        return await handle.evaluate(
            """(dialog, spec) => {
                const anchors = dialog.querySelectorAll('a[href]');
                const normalized = (spec.normalized || '').toLowerCase();
                const fallback = (spec.raw_lower || '').toLowerCase();
                const fallbackCompact = (spec.raw_lower_compact || '').toLowerCase();
                const displayLower = (spec.display_lower || '').toLowerCase();
                const displayCompact = (spec.display_compact || '').toLowerCase();
                for (const anchor of anchors) {
                    const href = (anchor.getAttribute('href') || '').toLowerCase();
                    const text = (anchor.textContent || '').toLowerCase();
                    const textCompact = text.replace(/\\s+/g, '');
                    if (normalized && href.includes(`/${normalized}`)) {
                        return true;
                    }
                    if (fallback && text.includes(fallback)) {
                        return true;
                    }
                    if (fallbackCompact && textCompact.includes(fallbackCompact)) {
                        return true;
                    }
                    if (displayLower && text.includes(displayLower)) {
                        return true;
                    }
                    if (displayCompact && textCompact.includes(displayCompact)) {
                        return true;
                    }
                }
                return false;
            }""",
            target_spec
        )
    except Exception:
        return False


async def inspect_profile_following(page, profile_url, target_specs):
    """단일 사용자 프로필에서 지정된 대상 팔로우 여부 확인"""
    default_status = {spec["label"]: "N/A" for spec in target_specs}
    if not profile_url:
        return default_status

    try:
        await page.goto(profile_url)
        await page.wait_for_timeout(random.uniform(1.2, 1.8) * 1000)
        await page.wait_for_selector("header", timeout=15000)
    except Exception as exc:
        print(f"      ⚠️  프로필 로드 실패: {exc}")
        return default_status

    try:
        dialog = await open_following_modal(page)
        if not dialog:
            print("      ⚠️  팔로잉 목록을 열 수 없습니다 (비공개 계정일 수 있음)")
            return default_status

        search_input = await wait_for_follow_search_input(page)
        if not search_input:
            print("      ⚠️  팔로우 검색창을 찾지 못했습니다")
            return default_status

        await search_input.click()
        results = default_status.copy()

        for spec in target_specs:
            try:
                query = spec.get("search_term") or spec["label"]
                await search_input.fill(query)
                await page.wait_for_timeout(random.uniform(1.0, 1.6) * 1000)

                found = await detect_follow_targets_in_dialog(dialog, spec)
                results[spec["label"]] = "O" if found else "X"
            except Exception as exc:
                print(f"        ⚠️  검색 실패 ({spec['label']}): {exc}")
                results[spec["label"]] = "N/A"
            finally:
                await search_input.fill("")
                await page.wait_for_timeout(400)

        try:
            await page.keyboard.press("Escape")
        except Exception:
            pass
        await page.wait_for_timeout(800)
        return results

    except Exception as exc:
        print(f"      ⚠️  팔로우 확인 실패: {exc}")
        return default_status


async def check_follow_status_for_comments(page, comments, target_specs):
    """댓글 목록 기준으로 각 사용자 팔로우 여부 확인"""
    if not target_specs:
        return comments

    profile_map = {}
    for comment in comments:
        username = comment.get("username", "").strip()
        if not username or username in profile_map:
            continue
        profile_url = build_profile_url(comment.get("profile_link"), username)
        if profile_url:
            profile_map[username] = profile_url

    if not profile_map:
        print("  ⚠️  팔로우 여부를 확인할 프로필 링크를 찾을 수 없습니다")
        return comments

    total = len(profile_map)
    print(f"\n{'═' * 70}")
    print(f"  🔎 팔로우 여부 확인 시작 (대상 {total}명)")
    print(f"{'═' * 70}")

    results_by_user = {}
    for idx, (username, profile_url) in enumerate(profile_map.items(), start=1):
        print(f"  [{idx}/{total}] @{username} 검사 중...", flush=True)
        statuses = await inspect_profile_following(page, profile_url, target_specs)
        summary = " | ".join([f"{label}:{statuses[label]}" for label in statuses])
        print(f"      → {summary}")
        results_by_user[username] = statuses

        if idx < total:
            await page.wait_for_timeout(random.uniform(1.8, 3.5) * 1000)

    print(f"\n  {'─' * 66}")
    print(f"  ✅ 팔로우 여부 확인 완료")
    print(f"  {'─' * 66}\n")

    for comment in comments:
        username = comment.get("username", "").strip()
        if username in results_by_user:
            comment["follow_status"] = results_by_user[username]

    return comments


# -----------------------
# 메인
# -----------------------
async def main():
    print(f"\n{'═' * 70}")
    print(f"  📸 Instagram 댓글 수집기")
    print(f"{'═' * 70}\n")

    # 필수 라이브러리 체크
    check_dependencies()

    # 로그인 방식 선택
    print(f"{'─' * 70}")
    print(f"  🔑 로그인 방식 선택:")
    print(f"  {'─' * 66}")
    print(f"      1. Username/Password 로그인 (간편)")
    print(f"      2. 쿠키 직접 입력")
    print(f"      3. .env 파일에서 쿠키 로드")
    print(f"  {'─' * 66}")
    login_mode = input("  선택 (1/2/3, 엔터 = 1): ").strip()

    use_login = False
    username = None
    password = None
    cookies = None

    if login_mode == '1':
        # Username/Password 방식
        use_login = True
        print(f"\n{'─' * 70}")
        print(f"  👤 Instagram 계정 정보 입력")
        print(f"{'─' * 70}")
        username = input("  Username: ").strip()
        if not username:
            print(f"  ❌ Username을 입력하세요")
            return

        import getpass
        password = getpass.getpass("  Password: ").strip()
        if not password:
            print(f"  ❌ Password를 입력하세요")
            return

    elif login_mode == '3':
        cookies = load_cookies_from_env()
        if not cookies:
            return
    else:
        cookies = get_cookies_from_user()
        if not cookies:
            print(f"\n  ❌ 쿠키 입력이 취소되었습니다")
            return

        # .env 저장 여부
        print(f"\n{'─' * 70}")
        save_choice = input("  💾 입력한 쿠키를 .env 파일로 저장하시겠습니까? (y/n, 엔터 = n): ").strip().lower()
        if save_choice in ['y', 'yes']:
            save_to_env_file(cookies)

    # 게시글 URL 입력
    print(f"\n{'─' * 70}")
    url = input("  🔗 Instagram 게시글 URL: ").strip()
    if not url:
        print(f"  ❌ URL을 입력하세요")
        return

    # 출력 파일명
    output = input(f"  📁 저장 파일명 (엔터 = comments.xlsx): ").strip() or "comments.xlsx"
    if not output.endswith(".xlsx") and not output.endswith(".csv"):
        output += ".xlsx"

    # 스크롤 모드 선택
    print(f"\n{'─' * 70}")
    print(f"  📜 스크롤 모드 선택:")
    print(f"  {'─' * 66}")
    print(f"      1. 자동 스크롤 (기본)")
    print(f"      2. 수동 스크롤 - 사용자가 직접 스크롤하면 실시간 수집")
    print(f"  {'─' * 66}")
    scroll_mode = input("  선택 (1/2, 엔터 = 1): ").strip()
    manual_scroll = scroll_mode == '2'

    # 최대 스크롤 (자동 모드만)
    max_scrolls = 100
    if not manual_scroll:
        max_scrolls_input = input(f"  🔢 최대 스크롤 (엔터 = 100): ").strip()
        max_scrolls = int(max_scrolls_input) if max_scrolls_input else 100

    # 헤드리스 모드
    headless_input = input(f"  👻 브라우저 숨김? (y/n, 엔터 = n): ").strip().lower()
    headless = headless_input in ['y', 'yes']

    # Playwright 실행
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=headless,
            args=[
                '--force-dark-mode=0',
                '--disable-blink-features=AutomationControlled'
            ]
        )

        context = await browser.new_context(
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            viewport={'width': 1280, 'height': 800},
            locale='ko-KR'
        )

        # Stealth 설정 (봇 감지 회피)
        await context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
        """)

        page = await context.new_page()

        try:
            if use_login:
                # Username/Password 로그인
                print(f"\n{'─' * 70}")
                print(f"  🔐 Instagram 로그인 중...")
                await page.goto('https://www.instagram.com/accounts/login/')
                await page.wait_for_timeout(3000)

                try:
                    await page.wait_for_selector('input[name="username"]', timeout=10000)

                    print(f"  ✍️  사용자명 입력 중...")
                    await page.fill('input[name="username"]', username)
                    await page.wait_for_timeout(1000)

                    print(f"  🔑 비밀번호 입력 중...")
                    await page.fill('input[name="password"]', password)
                    await page.wait_for_timeout(1000)

                    print(f"  🚀 로그인 버튼 클릭...")
                    await page.click('button[type="submit"]')
                    await page.wait_for_timeout(5000)

                    print(f"  ✅ 로그인 완료!")

                    # "나중에 하기" 버튼 처리
                    try:
                        not_now = page.locator('button:has-text("나중에 하기"), button:has-text("Not Now")')
                        if await not_now.count() > 0:
                            await not_now.first.click()
                            await page.wait_for_timeout(2000)
                    except:
                        pass

                    # 알림 버튼 처리
                    try:
                        not_now = page.locator('button:has-text("나중에 하기"), button:has-text("Not Now")')
                        if await not_now.count() > 0:
                            await not_now.first.click()
                            await page.wait_for_timeout(2000)
                    except:
                        pass

                except Exception as e:
                    print(f"  ❌ 로그인 실패: {e}")
                    return

                print(f"{'─' * 70}\n")

            else:
                # 쿠키 로드
                if not await load_cookies_to_context(context, cookies):
                    print(f"\n  ❌ 쿠키 로드 실패")
                    return

            # 댓글 수집
            if manual_scroll:
                comments = await extract_comments_manual(page, url)
            else:
                comments = await extract_comments_auto(page, url, max_scrolls)

            # 프로필 정보 수집 여부 확인
            if comments:
                unique_usernames = len(set(c.get('username', '') for c in comments if c.get('username')))

                print(f"\n{'─' * 70}")
                print(f"  👤 사용자 프로필 정보 수집 (선택사항)")
                print(f"  {'─' * 66}")
                print(f"      • 수집 가능한 정보: 게시물수, 팔로워수, 팔로잉수, 비공개계정여부")
                print(f"      • 대상 사용자: {unique_usernames}명 (중복 제거)")
                print(f"      • 예상 소요 시간: {unique_usernames * 4 / 60:.1f}분")
                print(f"  {'─' * 66}")
                print(f"\n  ⚠️  주의사항:")
                print(f"      • Instagram 서비스 약관 위반 가능성이 있습니다")
                print(f"      • 과도한 수집은 계정 정지로 이어질 수 있습니다")
                print(f"      • 중요한 계정 사용을 권장하지 않습니다")
                print(f"  {'─' * 66}\n")

                enrich_choice = input("  프로필 정보를 수집하시겠습니까? (y/n, 엔터 = n): ").strip().lower()

                if enrich_choice in ['y', 'yes']:
                    # 최대 수집 프로필 수 제한 옵션
                    max_profiles_input = input(f"  🔢 최대 프로필 수집 수 (엔터 = 전체 {unique_usernames}명): ").strip()
                    max_profiles = int(max_profiles_input) if max_profiles_input else None

                    # 프로필 정보 수집
                    comments = await enrich_comments_with_profile_data(page, comments, max_profiles)

                # 팔로우 확인 옵션
                follow_choice = input("\n  🎯 팔로우 여부를 추가로 확인하시겠습니까? (y/n, 엔터 = n): ").strip().lower()
                if follow_choice in ['y', 'yes']:
                    print("  💡 형식 예시: handle123, handle123|표시이름, handle123=>표시이름")
                    target_input = input("  🔎 확인할 계정을 입력하세요 (쉼표로 구분): ").strip()
                    target_specs = parse_follow_target_input(target_input)
                    if target_specs:
                        comments = await check_follow_status_for_comments(page, comments, target_specs)
                    else:
                        print("  ⚠️  유효한 팔로우 대상이 없어 건너뜁니다.")

            # 저장
            if comments:
                if output.endswith('.xlsx'):
                    save_to_excel(comments, output)
                else:
                    save_to_csv(comments, output)
                print(f"{'═' * 70}")
                print(f"  ✅ 모든 작업이 성공적으로 완료되었습니다!")
                print(f"{'═' * 70}\n")
            else:
                print(f"  ⚠️  추출된 데이터 없음")
                print(f"  💡 댓글이 없거나 DOM 구조 분석이 필요합니다")

        except KeyboardInterrupt:
            print(f"\n\n  ⏹️  사용자에 의해 중단됨")
        except Exception as e:
            print(f"\n  ❌ 오류 발생: {e}")
            import traceback
            traceback.print_exc()
        finally:
            print(f"\n  🔚 브라우저를 5초 후 종료합니다...")
            await page.wait_for_timeout(5000)
            await browser.close()


if __name__ == "__main__":
    asyncio.run(main())
