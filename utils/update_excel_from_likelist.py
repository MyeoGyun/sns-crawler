"""
likelist.md 파일을 읽어서 엑셀 파일의 미션 이행 여부를 업데이트하는 스크립트
사용법: python3 update_excel_from_likelist.py
"""

import io
import openpyxl
import msoffcrypto

# 설정
LIKELIST_FILE = '/Users/dinggyun/dev/retweet_check/likelist.md'
EXCEL_FILE_PATH = '/Users/dinggyun/dev/retweet_check/(잘함) 10월 자일리톨 수능 응원 이벤트 당첨자 리스트_1023.xlsx'
EXCEL_PASSWORD = 'xylitol_25'
OUTPUT_FILE_PATH = '/Users/dinggyun/dev/retweet_check/(잘함) 10월 자일리톨 수능 응원 이벤트 당첨자 리스트_1023_updated.xlsx'


def extract_all_entries_from_likelist(likelist_path):
    """likelist.md 파일에서 모든 항목을 추출합니다 (ID, 닉네임 모두 포함)."""
    print(f'\n{"="*60}')
    print('likelist.md에서 모든 항목 추출 중...')
    print(f'{"="*60}')

    liked_entries = set()

    try:
        with open(likelist_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        for line_num, line in enumerate(lines, 1):
            # 공백만 제거 (대소문자, @ 등은 그대로 유지)
            entry = line.strip()

            # 빈 줄이 아니면 추가
            if entry:
                liked_entries.add(entry)
                print(f'  {line_num}. {entry}')

    except Exception as e:
        print(f'❌ 파일 읽기 오류: {e}')
        return set()

    print(f'\n총 {len(liked_entries)}개 항목 추출 완료!')
    return liked_entries


def update_excel_file(liked_users, excel_path, password, output_path):
    """엑셀 파일의 미션 이행 여부를 업데이트합니다."""
    print(f'\n{"="*60}')
    print('엑셀 파일 업데이트 중...')
    print(f'{"="*60}')

    try:
        # 암호화된 파일 열기
        print('\n파일 복호화 중...')
        encrypted_file = open(excel_path, 'rb')
        decrypted = io.BytesIO()
        office_file = msoffcrypto.OfficeFile(encrypted_file)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
        encrypted_file.close()

        # 복호화된 파일 읽기
        decrypted.seek(0)
        wb = openpyxl.load_workbook(decrypted)
        ws = wb.active

        print(f'시트 이름: {ws.title}')
        print(f'총 행 수: {ws.max_row}')

        # 헤더 찾기 (3번째 행에 헤더가 있음)
        headers = [cell.value for cell in ws[3]]

        # ID 컬럼과 미션 이행 여부 컬럼 찾기
        id_col = None
        mission_col = None

        for idx, header in enumerate(headers, start=1):
            if header == 'ID':
                id_col = idx
            elif header == '미션 이행 여부':
                mission_col = idx

        if not id_col or not mission_col:
            print(f'❌ 필요한 컬럼을 찾을 수 없습니다. ID: {id_col}, 미션: {mission_col}')
            return

        print(f'\nID 컬럼: {id_col}번째 (열 {openpyxl.utils.get_column_letter(id_col)})')
        print(f'미션 이행 여부 컬럼: {mission_col}번째 (열 {openpyxl.utils.get_column_letter(mission_col)})')

        # 데이터 업데이트 (4번째 행부터 시작)
        updated_count = 0
        matched_count = 0
        matched_users = []
        unmatched_users = []

        print('\n미션 이행 여부 업데이트 중...')

        for row_idx in range(4, ws.max_row + 1):
            instagram_id = ws.cell(row=row_idx, column=id_col).value

            if instagram_id:
                # 원본 ID (공백만 제거, 정제 안함)
                original_id = str(instagram_id).strip()

                # 매칭 확인: 완전 일치 또는 부분 포함
                is_matched = False
                matched_entry = None

                # 1. 완전 일치 확인
                if original_id in liked_users:
                    is_matched = True
                    matched_entry = original_id
                else:
                    # 2. 부분 포함 확인 (liked_users의 각 항목이 original_id에 포함되는지)
                    for entry in liked_users:
                        if entry in original_id or original_id in entry:
                            is_matched = True
                            matched_entry = entry
                            break

                if is_matched:
                    ws.cell(row=row_idx, column=mission_col).value = 'O'
                    matched_count += 1
                    matched_users.append(original_id)
                    print(f'  ✓ 매칭: {original_id} (매칭된 항목: {matched_entry})')
                else:
                    ws.cell(row=row_idx, column=mission_col).value = 'X'
                    unmatched_users.append(original_id)

                updated_count += 1

        print(f'\n{"="*60}')
        print('업데이트 결과:')
        print(f'{"="*60}')
        print(f'  - 총 업데이트된 행: {updated_count}')
        print(f'  - 좋아요 매칭 (O): {matched_count}')
        print(f'  - 매칭 안됨 (X): {updated_count - matched_count}')

        if matched_users:
            print(f'\n✓ 매칭된 사용자 목록 ({len(matched_users)}명):')
            for idx, user in enumerate(sorted(matched_users), 1):
                print(f'  {idx}. {user}')

        if unmatched_users:
            print(f'\n✗ 매칭되지 않은 사용자 목록 ({len(unmatched_users)}명):')
            # 처음 50명만 출력
            for idx, user in enumerate(sorted(unmatched_users)[:50], 1):
                print(f'  {idx}. {user}')
            if len(unmatched_users) > 50:
                print(f'  ... 외 {len(unmatched_users) - 50}명 더 있음')

        # 파일 저장
        print(f'\n파일 저장 중: {output_path}')
        wb.save(output_path)
        wb.close()

        print(f'\n✓ 파일이 저장되었습니다!')
        print(f'  파일 경로: {output_path}')
        print(f'{"="*60}')

        # 결과를 .md 파일로 저장
        return matched_users, unmatched_users

    except Exception as e:
        print(f'\n❌ 엑셀 업데이트 오류: {e}')
        import traceback
        traceback.print_exc()
        return [], []


def save_results_to_md(matched_users, unmatched_users):
    """매칭 결과를 .md 파일로 저장합니다."""
    print(f'\n{"="*60}')
    print('결과를 .md 파일로 저장 중...')
    print(f'{"="*60}')

    try:
        # 매칭된 사용자 저장
        matched_file = '/Users/dinggyun/dev/retweet_check/매칭된_사용자.md'
        with open(matched_file, 'w', encoding='utf-8') as f:
            f.write('# 매칭된 사용자 목록\n\n')
            f.write(f'총 {len(matched_users)}명이 좋아요를 눌렀습니다.\n\n')
            f.write('---\n\n')
            for idx, user in enumerate(sorted(matched_users), 1):
                f.write(f'{idx}. {user}\n')

        print(f'\n✓ 매칭된 사용자 파일 저장 완료!')
        print(f'  파일: {matched_file}')
        print(f'  인원: {len(matched_users)}명')

        # 매칭되지 않은 사용자 저장
        unmatched_file = '/Users/dinggyun/dev/retweet_check/매칭되지_않은_사용자.md'
        with open(unmatched_file, 'w', encoding='utf-8') as f:
            f.write('# 매칭되지 않은 사용자 목록\n\n')
            f.write(f'총 {len(unmatched_users)}명이 좋아요를 누르지 않았습니다.\n\n')
            f.write('---\n\n')
            for idx, user in enumerate(sorted(unmatched_users), 1):
                f.write(f'{idx}. {user}\n')

        print(f'\n✓ 매칭되지 않은 사용자 파일 저장 완료!')
        print(f'  파일: {unmatched_file}')
        print(f'  인원: {len(unmatched_users)}명')

        print(f'\n{"="*60}')

    except Exception as e:
        print(f'\n❌ .md 파일 저장 오류: {e}')
        import traceback
        traceback.print_exc()


def main():
    """메인 함수"""
    print('\n자일리톨 이벤트 좋아요 확인 프로그램')
    print('likelist.md 파일 기반')

    # 1. likelist.md에서 모든 항목 추출
    liked_users = extract_all_entries_from_likelist(LIKELIST_FILE)

    if not liked_users:
        print('\n❌ Instagram ID를 추출하지 못했습니다.')
        return

    print(f'\n추출된 ID 샘플 (처음 10개):')
    for idx, user in enumerate(sorted(liked_users)[:10], 1):
        print(f'  {idx}. {user}')

    # 2. 엑셀 파일 업데이트
    matched_users, unmatched_users = update_excel_file(liked_users, EXCEL_FILE_PATH, EXCEL_PASSWORD, OUTPUT_FILE_PATH)

    # 3. 결과를 .md 파일로 저장
    if matched_users or unmatched_users:
        save_results_to_md(matched_users, unmatched_users)

    print('\n모든 작업이 완료되었습니다!')


if __name__ == '__main__':
    main()
