#!/usr/bin/env python3
# quote_extractor_cookie_input.py
# CLI에서 직접 쿠키를 입력받아 자동 로그인
#
# 사용법:
# python quote_extractor_cookie_input.py

import os
import time
import csv
import getpass
import random
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# undetected-chromedriver 시도 (봇 감지 회피)
try:
    import undetected_chromedriver as uc
    USE_UNDETECTED = True
except ImportError:
    USE_UNDETECTED = False
    print("⚠️  [경고] undetected-chromedriver가 설치되지 않았습니다")
    print("    pip install undetected-chromedriver 로 설치하면 봇 감지 회피 성능이 향상됩니다\n")

try:
    from webdriver_manager.chrome import ChromeDriverManager
    AUTO_DRIVER = True
except ImportError:
    AUTO_DRIVER = False

# -----------------------
# 설정값
# -----------------------
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_FILE = os.path.join(BASE_DIR, ".env")

# 봇 회피 설정
DELAY_INCREASE_THRESHOLD = 60  # 딜레이 증가 시작 트윗 수
DELAY_INCREASE_STEP = 10       # 매 10개마다 증가
DELAY_INCREASE_AMOUNT = 2      # 증가량 (초)
MAX_DELAY = 20                 # 최대 딜레이 (초)

REST_INTERVALS = [60, 120, 180, 240, 300]  # 휴식 지점
REST_MIN = 30                  # 최소 휴식 시간 (초)
REST_MAX = 60                  # 최대 휴식 시간 (초)

# -----------------------
# 랜덤 딜레이 헬퍼 함수
# -----------------------
def random_delay(min_sec=8, max_sec=12, description=""):
    """
    랜덤한 지연 시간을 생성하여 봇 감지를 회피합니다.

    Args:
        min_sec: 최소 대기 시간 (초)
        max_sec: 최대 대기 시간 (초)
        description: 대기 이유 설명
    """
    delay = random.uniform(min_sec, max_sec)
    if description:
        print(f"  ⏱️  {description} ({delay:.1f}초)")
    time.sleep(delay)
    return delay

def get_adaptive_delay(tweet_count):
    """
    트윗 수에 따라 딜레이를 점진적으로 증가시킵니다.

    Args:
        tweet_count: 현재까지 수집한 트윗 수

    Returns:
        tuple: (min_delay, max_delay)
    """
    if tweet_count <= DELAY_INCREASE_THRESHOLD:
        return (8, 12)  # 기본 딜레이

    # 60개 이후 매 10개마다 딜레이 증가
    extra = ((tweet_count - DELAY_INCREASE_THRESHOLD) // DELAY_INCREASE_STEP) * DELAY_INCREASE_AMOUNT
    min_delay = min(8 + extra, MAX_DELAY)
    max_delay = min(12 + extra, MAX_DELAY + 2)

    return (min_delay, max_delay)

def take_rest_if_needed(tweet_count):
    """
    특정 트윗 개수마다 휴식 시간을 부여합니다.

    Args:
        tweet_count: 현재까지 수집한 트윗 수

    Returns:
        bool: 휴식을 취했으면 True
    """
    if tweet_count in REST_INTERVALS:
        rest_time = random.uniform(REST_MIN, REST_MAX)
        print(f"\n{'─' * 70}")
        print(f"  🛑 [휴식 모드 - 봇 감지 회피]")
        print(f"  {'─' * 66}")
        print(f"      📊 현재 수집: {tweet_count}개")
        print(f"      ⏰ 휴식 시간: {rest_time:.0f}초")
        print(f"      💡 인간 행동 패턴을 모방하여 봇 감지를 회피합니다")
        print(f"  {'─' * 66}")
        print(f"      대기 중", end="", flush=True)

        # 진행 표시
        for i in range(int(rest_time)):
            time.sleep(1)
            if i % 5 == 0:
                print(".", end="", flush=True)

        print(" ✅ 완료!")
        print(f"{'─' * 70}\n")
        return True
    return False

# -----------------------
# 쿠키 입력받기
# -----------------------
def get_cookies_from_user():
    """
    사용자로부터 쿠키 값을 직접 입력받기

    Returns:
        dict: 쿠키 딕셔너리 또는 None
    """
    print("\n" + "=" * 60)
    print("   쿠키 입력")
    print("=" * 60)
    print()
    print("Chrome에서 쿠키를 추출하는 방법:")
    print("1. Chrome 브라우저에서 https://x.com 로그인")
    print("2. F12 (개발자 도구) 열기")
    print("3. Application 탭 → Cookies → https://x.com")
    print("4. 아래 쿠키 값들을 복사하여 붙여넣기")
    print()
    print("자세한 가이드: README_COOKIES.md 참고")
    print("=" * 60)
    print()

    cookies = {}

    # 필수 쿠키 입력
    print("📌 필수 쿠키 (반드시 필요):")
    print()

    # AUTH_TOKEN
    print("1. auth_token 쿠키 값을 입력하세요:")
    print("   (Chrome에서 auth_token 값을 복사하여 붙여넣기)")
    auth_token = input("AUTH_TOKEN: ").strip()

    if not auth_token:
        print("\n[오류] AUTH_TOKEN은 필수입니다!")
        return None

    cookies["AUTH_TOKEN"] = auth_token
    print("✅ AUTH_TOKEN 입력 완료")
    print()

    # CT0
    print("2. ct0 쿠키 값을 입력하세요:")
    print("   (Chrome에서 ct0 값을 복사하여 붙여넣기)")
    ct0 = input("CT0: ").strip()

    if not ct0:
        print("\n[오류] CT0은 필수입니다!")
        return None

    cookies["CT0"] = ct0
    print("✅ CT0 입력 완료")
    print()

    # 선택 쿠키
    print("📝 선택 쿠키 (없으면 엔터):")
    print()

    print("3. twid 쿠키 값 (선택, 엔터로 건너뛰기):")
    twid = input("TWID: ").strip()
    if twid:
        cookies["TWID"] = twid
        print("✅ TWID 입력 완료")
    else:
        print("⏭️  TWID 건너뜀")
    print()

    print("4. guest_id 쿠키 값 (선택, 엔터로 건너뛰기):")
    guest_id = input("GUEST_ID: ").strip()
    if guest_id:
        cookies["GUEST_ID"] = guest_id
        print("✅ GUEST_ID 입력 완료")
    else:
        print("⏭️  GUEST_ID 건너뜀")
    print()

    return cookies

# -----------------------
# .env 파일로 저장
# -----------------------
def save_to_env_file(cookies):
    """
    입력받은 쿠키를 .env 파일로 저장

    Args:
        cookies: 쿠키 딕셔너리

    Returns:
        bool: 저장 성공 여부
    """
    try:
        with open(ENV_FILE, "w", encoding="utf-8") as f:
            f.write("# Twitter/X 쿠키 설정\n")
            f.write("# 이 파일을 다른 사람과 공유하지 마세요!\n")
            f.write("\n")
            f.write("# 필수 쿠키\n")
            f.write(f"AUTH_TOKEN={cookies.get('AUTH_TOKEN', '')}\n")
            f.write(f"CT0={cookies.get('CT0', '')}\n")
            f.write("\n")
            f.write("# 선택 쿠키\n")
            f.write(f"TWID={cookies.get('TWID', '')}\n")
            f.write(f"GUEST_ID={cookies.get('GUEST_ID', '')}\n")
            f.write("\n")
            f.write("# 전체 쿠키 JSON (사용 안 함)\n")
            f.write("COOKIES_JSON=\n")

        print(f"\n[성공] 쿠키를 .env 파일로 저장했습니다: {ENV_FILE}")
        print("[안내] 다음 실행 시 quote_extractor_env.py로 자동 로그인 가능")
        return True

    except Exception as e:
        print(f"\n[경고] .env 저장 실패: {e}")
        return False

# -----------------------
# 쿠키 로드 및 적용
# -----------------------
def load_cookies_to_driver(driver, cookies):
    """
    입력받은 쿠키를 Selenium에 적용

    Args:
        driver: Selenium WebDriver
        cookies: 쿠키 딕셔너리

    Returns:
        bool: 성공 여부
    """
    try:
        # Twitter 홈페이지로 먼저 이동 (쿠키 도메인 일치 필요)
        print(f"\n{'─' * 70}")
        print(f"  🌐 Twitter 접속 중...")
        driver.get("https://x.com")
        random_delay(3, 6, "페이지 초기 로딩")

        print(f"  🔑 쿠키 로드 중...")

        # 필수 쿠키
        required_cookies = {
            "auth_token": cookies.get("AUTH_TOKEN"),
            "ct0": cookies.get("CT0"),
        }

        # 선택 쿠키
        optional_cookies = {
            "twid": cookies.get("TWID"),
            "guest_id": cookies.get("GUEST_ID"),
        }

        # 필수 쿠키 확인
        missing = [k for k, v in required_cookies.items() if not v]
        if missing:
            print(f"  ❌ 필수 쿠키 누락: {', '.join(missing)}")
            return False

        # 쿠키 추가
        cookies_added = 0
        for name, value in {**required_cookies, **optional_cookies}.items():
            if value:
                try:
                    driver.add_cookie({
                        "name": name,
                        "value": value,
                        "domain": ".x.com",
                        "path": "/",
                    })
                    cookies_added += 1
                except Exception as e:
                    print(f"  ⚠️  쿠키 추가 실패 ({name}): {e}")

        print(f"  ✅ {cookies_added}개 쿠키 로드 완료")

        # 쿠키 적용 확인
        print(f"  🔐 세션 검증 중...")
        driver.get("https://x.com/home")
        random_delay(5, 8, "로그인 세션 검증")

        # 로그인 상태 확인
        if is_logged_in(driver):
            print(f"  ✅ 쿠키로 로그인 성공!")
            print(f"{'─' * 70}\n")
            return True
        else:
            print(f"  ❌ 쿠키가 유효하지 않습니다")
            print(f"  ℹ️  Chrome에서 다시 로그인하여 새 쿠키를 가져오세요")
            print(f"      (쿠키 만료 또는 값이 잘못되었을 수 있습니다)")
            print(f"{'─' * 70}\n")
            return False

    except Exception as e:
        print(f"  ❌ 쿠키 로드 실패: {e}")
        print(f"{'─' * 70}\n")
        return False

# -----------------------
# 로그인 상태 확인
# -----------------------
def is_logged_in(driver, timeout=5):
    """현재 로그인 상태인지 확인"""
    try:
        if "x.com/home" in driver.current_url or "twitter.com/home" in driver.current_url:
            return True

        home_indicators = [
            (By.CSS_SELECTOR, '[data-testid="AppTabBar_Home_Link"]'),
            (By.CSS_SELECTOR, '[data-testid="SideNav_AccountSwitcher_Button"]'),
        ]

        for by, selector in home_indicators:
            try:
                element = WebDriverWait(driver, timeout).until(
                    EC.presence_of_element_located((by, selector))
                )
                if element.is_displayed():
                    return True
            except:
                continue

        return False
    except:
        return False

# -----------------------
# Selenium 드라이버
# -----------------------
def make_driver(headless=True):
    """
    Selenium WebDriver 생성 (가능하면 undetected-chromedriver 사용)

    Args:
        headless: 헤드리스 모드 사용 여부

    Returns:
        WebDriver 인스턴스
    """
    if USE_UNDETECTED:
        # undetected-chromedriver 사용 (봇 감지 회피)
        print(f"  🛡️  undetected-chromedriver 사용 (고급 봇 회피)")

        options = uc.ChromeOptions()
        if headless:
            options.add_argument("--headless=new")

        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--window-size=1200,2000")

        # 라이트 모드 강제 설정
        options.add_argument("--force-dark-mode=0")
        options.add_argument("--disable-features=WebUIDarkMode")
        options.add_argument("--disable-features=DarkMode")
        options.add_argument("--force-color-profile=srgb")

        # Chrome preferences
        prefs = {
            "profile.default_content_setting_values.automatic_downloads": 1,
            "download.prompt_for_download": False,
            "profile.default_content_setting_values.theme": 0,
            "profile.managed_default_content_settings.theme": 0,
        }
        options.add_experimental_option("prefs", prefs)

        # Local state
        options.add_experimental_option("localState", {
            "browser": {
                "enabled_labs_experiments": [
                    "disable-webui-dark-mode@1",
                    "disable-dark-mode@1"
                ]
            }
        })

        try:
            driver = uc.Chrome(options=options, version_main=131)
        except:
            # 버전 자동 감지 실패 시
            driver = uc.Chrome(options=options)

    else:
        # 일반 Selenium WebDriver 사용
        print(f"  ⚙️  일반 Selenium WebDriver 사용")

        opts = Options()
        if headless:
            opts.add_argument("--headless=new")
        opts.add_argument(f"user-agent={USER_AGENT}")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--window-size=1200,2000")

        # 라이트 모드 강제 설정
        opts.add_argument("--force-dark-mode=0")
        opts.add_argument("--disable-features=WebUIDarkMode")
        opts.add_argument("--disable-features=DarkMode")
        opts.add_argument("--force-color-profile=srgb")

        # Chrome preferences
        prefs = {
            "profile.default_content_setting_values.automatic_downloads": 1,
            "download.prompt_for_download": False,
            "profile.default_content_setting_values.theme": 0,
            "profile.managed_default_content_settings.theme": 0,
        }
        opts.add_experimental_option("prefs", prefs)

        # Local state
        opts.add_experimental_option("localState", {
            "browser": {
                "enabled_labs_experiments": [
                    "disable-webui-dark-mode@1",
                    "disable-dark-mode@1"
                ]
            }
        })

        if AUTO_DRIVER:
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=opts)
        else:
            driver = webdriver.Chrome(options=opts)

    driver.set_page_load_timeout(60)
    return driver

# -----------------------
# 파싱
# -----------------------
def parse_quote_tweet(article, driver=None, screenshot_dir="screenshots"):
    """
    트윗 파싱 및 스크린샷 캡처 (독립적 필드 파싱)
    각 필드를 독립적으로 처리하여 일부 필드 실패해도 나머지 데이터는 수집
    """
    data = {
        "status_id": "", "url": "", "author_handle": "", "text": "",
        "hashtags": "", "time_iso_utc": "", "has_media": "",
        "media_urls": "", "is_quote": "FALSE",
        "quote_status_id": "인용X", "quote_time_iso_utc": "인용X",
        "reply_count": "0", "retweet_count": "0", "like_count": "0",
        "screenshot_path": "",
    }

    # 1) status_id 및 url - time 요소의 부모 링크에서 추출 (가장 정확)
    try:
        time_element = article.find_element(By.CSS_SELECTOR, "time[datetime]")
        # time 요소의 부모 a 태그가 이 트윗의 정확한 링크
        parent_link = time_element.find_element(By.XPATH, "..")
        href = parent_link.get_attribute("href")
        if href and "/status/" in href:
            parts = href.split("/status/")
            if len(parts) == 2:
                status_id = parts[1].split("?")[0]
                data["status_id"] = status_id
                # URL 정규화
                username_part = parts[0].replace('https://x.com', '').replace('https://twitter.com', '')
                data["url"] = f"https://x.com{username_part}/status/{status_id}"
    except Exception as e:
        # 대안: 일반 status 링크 사용
        try:
            time_link = article.find_element(By.CSS_SELECTOR, "a[href*='/status/']")
            href = time_link.get_attribute("href")
            if href and "/status/" in href:
                parts = href.split("/status/")
                if len(parts) == 2:
                    status_id = parts[1].split("?")[0]
                    data["status_id"] = status_id
                    username_part = parts[0].replace('https://x.com', '').replace('https://twitter.com', '')
                    data["url"] = f"https://x.com{username_part}/status/{status_id}"
        except:
            pass

    # status_id가 없으면 이 트윗은 무효
    if not data["status_id"]:
        return None

    # 2) author_handle
    try:
        profile_link = article.find_element(By.CSS_SELECTOR, "a[href^='/'][role='link']")
        profile_href = profile_link.get_attribute("href")
        if profile_href:
            username = profile_href.split("?")[0].split("/")[-1]
            if username and not username.startswith("status"):
                data["author_handle"] = f"@{username}" if not username.startswith("@") else username
    except:
        # 대안: span 텍스트에서 @ 찾기
        try:
            spans = article.find_elements(By.CSS_SELECTOR, "span")
            for span in spans:
                text = span.text.strip()
                if text.startswith("@"):
                    data["author_handle"] = text
                    break
        except:
            pass

    # 3) text 및 hashtags
    try:
        tweet_text_div = article.find_element(By.CSS_SELECTOR, "[data-testid='tweetText']")
        data["text"] = tweet_text_div.text.strip()

        # 해시태그 추출
        try:
            hashtag_links = tweet_text_div.find_elements(By.CSS_SELECTOR, "a[href*='/hashtag/']")
            hashtags = [link.text.strip() for link in hashtag_links if link.text.strip()]
            data["hashtags"] = ", ".join(hashtags) if hashtags else ""
        except:
            data["hashtags"] = ""
    except:
        # 텍스트 없는 트윗도 허용 (이미지 전용 등)
        data["text"] = ""
        data["hashtags"] = ""

    # 4) time_iso_utc
    try:
        time_element = article.find_element(By.CSS_SELECTOR, "time[datetime]")
        datetime_attr = time_element.get_attribute("datetime")
        if datetime_attr:
            data["time_iso_utc"] = datetime_attr
    except:
        pass

    # 5) media (이미지 및 비디오)
    media_urls = []
    try:
        # 이미지
        images = article.find_elements(By.CSS_SELECTOR, "img[src*='pbs.twimg.com']")
        for img in images:
            src = img.get_attribute("src")
            if src and "profile_images" not in src:
                media_urls.append(src)
    except:
        pass

    try:
        # 비디오
        videos = article.find_elements(By.CSS_SELECTOR, "video source")
        for video in videos:
            src = video.get_attribute("src")
            if src:
                media_urls.append(src)
    except:
        pass

    if media_urls:
        data["has_media"] = "TRUE"
        data["media_urls"] = ", ".join(media_urls)
    else:
        data["has_media"] = "FALSE"
        data["media_urls"] = "인용X"

    # 6) reply_count, retweet_count, like_count
    try:
        reply_btn = article.find_element(By.CSS_SELECTOR, "button[data-testid='reply']")
        reply_aria = reply_btn.get_attribute("aria-label") or ""
        for part in reply_aria.split():
            if part.isdigit():
                data["reply_count"] = part
                break
    except:
        pass

    try:
        retweet_btn = article.find_element(By.CSS_SELECTOR, "button[data-testid='retweet']")
        retweet_aria = retweet_btn.get_attribute("aria-label") or ""
        for part in retweet_aria.split():
            if part.isdigit():
                data["retweet_count"] = part
                break
    except:
        pass

    try:
        like_btn = article.find_element(By.CSS_SELECTOR, "button[data-testid='like']")
        like_aria = like_btn.get_attribute("aria-label") or ""
        for part in like_aria.split():
            if part.isdigit():
                data["like_count"] = part
                break
    except:
        pass

    # 7) 스크린샷 캡처
    if driver and data["status_id"]:
        try:
            os.makedirs(screenshot_dir, exist_ok=True)
            screenshot_filename = f"@스크린샷_tweet_{data['status_id']}.png"
            screenshot_path = os.path.join(screenshot_dir, screenshot_filename)

            # 부드럽게 스크롤 (메인 스크롤 방해 최소화)
            driver.execute_script("""
                arguments[0].scrollIntoView({block: 'nearest', behavior: 'smooth'});
            """, article)
            time.sleep(0.2)

            article.screenshot(screenshot_path)
            data["screenshot_path"] = screenshot_path
        except:
            data["screenshot_path"] = ""

    return data

# -----------------------
# 추출
# -----------------------
def extract_quote_tweets(driver, quotes_url, max_scrolls=None, max_tweets=None, capture_screenshots=True):
    """
    인용글 추출 (봇 회피 기능 포함)

    Args:
        driver: Selenium WebDriver
        quotes_url: 인용글 URL
        max_scrolls: 최대 스크롤 횟수
        max_tweets: 최대 수집 개수
        capture_screenshots: 스크린샷 캡처 여부

    Returns:
        list: 추출된 트윗 데이터
    """
    print(f"\n{'─' * 70}")
    print(f"  📄 인용글 페이지 접근 중...")
    if max_tweets:
        print(f"  🎯 목표: 최대 {max_tweets}개 수집")
    print(f"{'─' * 70}")
    driver.get(quotes_url)
    random_delay(5, 8, "페이지 로딩")

    # 강제 라이트 모드 적용 (JavaScript로 직접 설정)
    try:
        print(f"  🌞 라이트 모드 강제 적용 중...")
        driver.execute_script("""
            // 1. HTML 요소의 다크 모드 속성 제거
            document.documentElement.removeAttribute('data-color-mode');
            document.documentElement.removeAttribute('data-theme');
            document.documentElement.style.colorScheme = 'light';

            // 2. body의 다크 모드 클래스 제거
            document.body.classList.remove('dark');
            document.body.classList.add('light');

            // 3. CSS 변수로 강제 라이트 모드 색상 적용
            document.documentElement.style.setProperty('--background-color', '#ffffff');
            document.documentElement.style.setProperty('--color-background', '#ffffff');

            // 4. localStorage에 라이트 모드 설정 저장
            try {
                localStorage.setItem('theme', 'light');
                localStorage.setItem('nightMode', 'false');
            } catch(e) {}
        """)
        print(f"  ✅ 라이트 모드 적용 완료")
    except Exception as e:
        print(f"  ⚠️  라이트 모드 적용 실패 (무시하고 계속): {e}")

    all_tweets = []
    seen_ids = set()
    scrolls = 0
    no_new_data = 0
    consecutive_no_dom_change = 0

    print(f"\n{'═' * 70}")
    print(f"  🔄 스크롤 및 데이터 수집 시작")
    print(f"{'═' * 70}\n")

    while True:
        # 현재 DOM에 있는 article 개수 기록 (무한 스크롤 감지용)
        try:
            articles_before_scroll = len(driver.find_elements(By.CSS_SELECTOR, "article[data-testid='tweet']"))
        except:
            articles_before_scroll = 0

        # 현재 페이지의 모든 트윗 파싱
        try:
            articles = driver.find_elements(By.CSS_SELECTOR, "article[data-testid='tweet']")
        except:
            print("  ⚠️  article 요소를 찾을 수 없습니다")
            break

        new = 0
        for article in articles:
            # 스크린샷 캡처가 활성화된 경우에만 driver 전달
            tweet = parse_quote_tweet(article, driver=driver if capture_screenshots else None)
            if tweet and tweet["status_id"] and tweet["status_id"] not in seen_ids:
                seen_ids.add(tweet["status_id"])
                all_tweets.append(tweet)
                new += 1

                # 각 트윗 추출 시 즉시 로그 출력
                text_preview = tweet["text"][:45] + "..." if len(tweet["text"]) > 45 else tweet["text"]
                author = tweet['author_handle'][:18]
                media_icon = "📎" if tweet["has_media"] == "TRUE" else "  "
                print(f"  {media_icon} #{len(all_tweets):3d}  {author:18s}  {text_preview}")

                # 🔥 최대 개수 체크
                if max_tweets and len(all_tweets) >= max_tweets:
                    print(f"\n{'─' * 70}")
                    print(f"  ✅ 최대 수집 개수 ({max_tweets}개) 도달 - 수집 종료")
                    print(f"{'═' * 70}\n")
                    return all_tweets

        # 스크롤 정보 표시
        scrolls += 1
        progress = f"({scrolls}/{max_scrolls})" if max_scrolls else ""

        print(f"\n  {'─' * 66}")
        print(f"  📊 스크롤 #{scrolls:02d} {progress:8s}  |  DOM: {len(articles):3d}개  |  신규: {new:2d}개  |  총: {len(all_tweets):3d}개")

        # 신규 데이터 없음 카운터
        if new == 0:
            no_new_data += 1
        else:
            no_new_data = 0

        # 최대 스크롤 도달 확인
        if max_scrolls and scrolls >= max_scrolls:
            print(f"  ⏹️  최대 스크롤 횟수 ({max_scrolls}회) 도달 - 수집 종료")
            break

        # 부드러운 스크롤 실행 (봇 감지 회피)
        last_h = driver.execute_script("return document.body.scrollHeight")
        current_position = driver.execute_script("return window.pageYOffset")

        # 현재 위치에서 조금씩 스크롤 (사람처럼)
        scroll_steps = random.randint(3, 6)  # 3~6단계로 나눠서 스크롤
        target_position = last_h
        step_size = (target_position - current_position) / scroll_steps

        for step in range(scroll_steps):
            next_position = current_position + (step_size * (step + 1))
            # 약간의 오버슈팅/언더슈팅으로 자연스러움 추가
            noise = random.uniform(-50, 50)
            next_position = max(0, next_position + noise)

            driver.execute_script(f"window.scrollTo({{top: {int(next_position)}, behavior: 'smooth'}});")
            time.sleep(random.uniform(0.1, 0.3))  # 각 스텝마다 짧은 대기

        # 🔥 봇 회피: 적응형 딜레이
        min_delay, max_delay = get_adaptive_delay(len(all_tweets))
        delay_desc = "트위터 API 대기"
        if len(all_tweets) > DELAY_INCREASE_THRESHOLD:
            delay_desc = f"트위터 API 대기 (적응형: {len(all_tweets)}개 수집)"
        random_delay(min_delay, max_delay, delay_desc)

        # DOM 변화 감지: 새로운 article이 추가되었는지 확인
        articles_after_scroll = len(driver.find_elements(By.CSS_SELECTOR, "article[data-testid='tweet']"))
        dom_changed = articles_after_scroll > articles_before_scroll

        if dom_changed:
            delta = articles_after_scroll - articles_before_scroll
            print(f"  ✅ DOM 업데이트: {articles_before_scroll} → {articles_after_scroll} (+{delta})")
            consecutive_no_dom_change = 0
        else:
            consecutive_no_dom_change += 1
            print(f"  ⏳ DOM 변화 없음 (연속 {consecutive_no_dom_change}회), 추가 대기 중...")

            # 🔥 "Show more" 버튼 찾아서 클릭 시도
            show_more_clicked = False
            try:
                # 다양한 "Show more" 버튼 선택자
                show_more_selectors = [
                    "//div[@role='button' and contains(., 'Show more')]",
                    "//div[@role='button' and contains(., 'more replies')]",
                    "//span[contains(text(), 'Show')]",
                    "//span[contains(text(), 'more')]",
                ]

                for selector in show_more_selectors:
                    try:
                        buttons = driver.find_elements(By.XPATH, selector)
                        for btn in buttons:
                            if btn.is_displayed():
                                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                                time.sleep(0.5)
                                btn.click()
                                print(f"  🔘 'Show more' 버튼 클릭 완료")
                                show_more_clicked = True
                                random_delay(3, 5, "Show more 로딩 대기")
                                break
                    except:
                        continue

                    if show_more_clicked:
                        break
            except:
                pass

            # Show more 클릭 후 또는 추가 대기
            if not show_more_clicked:
                random_delay(8, 12, "DOM 업데이트 대기")

            # 재확인
            articles_recheck = len(driver.find_elements(By.CSS_SELECTOR, "article[data-testid='tweet']"))
            if articles_recheck > articles_after_scroll:
                delta = articles_recheck - articles_after_scroll
                print(f"  ✅ 추가 대기 후 증가: {articles_after_scroll} → {articles_recheck} (+{delta})")
                consecutive_no_dom_change = 0
            else:
                # 페이지 높이도 확인
                new_h = driver.execute_script("return document.body.scrollHeight")
                if new_h == last_h:
                    # DOM 변화 없음 + 신규 데이터 없음 = 종료 조건 (완화됨)
                    if consecutive_no_dom_change >= 5 and no_new_data >= 5:
                        # 마지막 확인: 한 번 더 스크롤 시도
                        print(f"  🔍 페이지 끝 최종 확인 중...")
                        driver.execute_script("window.scrollBy(0, 500);")
                        random_delay(5, 8, "최종 확인 대기")

                        final_article_count = len(driver.find_elements(By.CSS_SELECTOR, "article[data-testid='tweet']"))
                        if final_article_count == articles_recheck:
                            print(f"\n  {'─' * 66}")
                            print(f"  🏁 수집 종료 조건 충족")
                            print(f"      • DOM 무변화: {consecutive_no_dom_change}회 연속")
                            print(f"      • 신규 데이터 없음: {no_new_data}회 연속")
                            print(f"      • 최종 확인 완료")
                            break
                        else:
                            print(f"  ✅ 최종 확인 후 추가 트윗 발견! 수집 계속...")
                            consecutive_no_dom_change = 0

                    # DOM만 변화 없고 신규 데이터는 있었다면 조금 더 기다림
                    if consecutive_no_dom_change >= 7:
                        print(f"\n  🏁 연속 7회 DOM 변화 없음 - 페이지 끝으로 판단")
                        break

        # 🔥 봇 회피: 휴식 시간 체크 (스크롤 후에 체크)
        take_rest_if_needed(len(all_tweets))

    print(f"\n{'═' * 70}")
    print(f"  ✅ 수집 완료!")
    print(f"  {'─' * 66}")
    print(f"      📝 총 수집 트윗:  {len(all_tweets)}개")
    print(f"      🔄 총 스크롤 횟수: {scrolls}회")
    final_article_count = len(driver.find_elements(By.CSS_SELECTOR, 'article[data-testid="tweet"]'))
    print(f"      📊 최종 DOM 개수:  {final_article_count}개")
    print(f"{'═' * 70}\n")
    return all_tweets

# -----------------------
# 엑셀 저장 (이미지 포함)
# -----------------------
def save_to_excel(tweets, output_path):
    """엑셀 파일 저장 (스크린샷 이미지 포함)"""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("  ⚠️  openpyxl 라이브러리가 필요합니다: pip install openpyxl Pillow")
        return

    if not tweets:
        print("  ⚠️  저장할 데이터 없음")
        return

    print(f"\n{'─' * 70}")
    print(f"  📊 엑셀 파일 생성 중...")
    print(f"{'─' * 70}")

    wb = Workbook()
    ws = wb.active
    ws.title = "인용글 목록"

    # 헤더 작성
    headers = ["스크린샷", "등록일자", "게시물번호", "좋아요수", "리트윗수", "댓글수",
               "내용", "해시태그", "게시물URL", "작성자URL", "작성자아이디"]
    ws.append(headers)

    # 헤더 스타일 적용
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 열 너비 설정
    ws.column_dimensions['A'].width = 35  # 스크린샷
    ws.column_dimensions['B'].width = 20  # 등록일자
    ws.column_dimensions['C'].width = 18  # 게시물번호
    ws.column_dimensions['D'].width = 10  # 좋아요수
    ws.column_dimensions['E'].width = 10  # 리트윗수
    ws.column_dimensions['F'].width = 10  # 댓글수
    ws.column_dimensions['G'].width = 50  # 내용
    ws.column_dimensions['H'].width = 30  # 해시태그
    ws.column_dimensions['I'].width = 40  # 게시물URL
    ws.column_dimensions['J'].width = 30  # 작성자URL
    ws.column_dimensions['K'].width = 20  # 작성자아이디

    # 데이터 행 추가
    for idx, tweet in enumerate(tweets, start=2):
        row_data = [
            "",  # 스크린샷 열 (이미지로 채움)
            tweet.get("time_iso_utc", ""),
            tweet.get("status_id", ""),
            tweet.get("like_count", "0"),
            tweet.get("retweet_count", "0"),
            tweet.get("reply_count", "0"),
            tweet.get("text", ""),
            tweet.get("hashtags", ""),
            tweet.get("url", ""),
            tweet.get("url", "").rsplit("/status/", 1)[0] if tweet.get("url") else "",
            tweet.get("author_handle", ""),
        ]
        ws.append(row_data)

        # 텍스트 셀 정렬
        for col_idx in range(2, 12):  # B열부터 K열까지
            cell = ws.cell(row=idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        # 스크린샷 이미지 삽입
        screenshot_path = tweet.get("screenshot_path", "")
        if screenshot_path and os.path.exists(screenshot_path):
            try:
                img = XLImage(screenshot_path)

                # 이미지 크기 조정 (높이 250px로 통일)
                target_height = 250
                aspect_ratio = img.width / img.height
                img.height = target_height
                img.width = int(target_height * aspect_ratio)

                # A열에 이미지 추가
                cell_address = f"A{idx}"
                ws.add_image(img, cell_address)

                # 행 높이 조정 (포인트 단위, 1px ≈ 0.75pt)
                ws.row_dimensions[idx].height = target_height * 0.75

                if idx % 10 == 0:
                    print(f"  📸 {idx-1}개 이미지 삽입 완료...")

            except Exception as e:
                print(f"  ⚠️  이미지 삽입 실패 (행 {idx}): {e}")
        else:
            # 이미지 없으면 행 높이 기본값
            ws.row_dimensions[idx].height = 20

    # 첫 행 고정 (헤더)
    ws.freeze_panes = "A2"

    # 파일 저장
    wb.save(output_path)

    print(f"\n{'─' * 70}")
    print(f"  💾 엑셀 파일 저장 완료!")
    print(f"  {'─' * 66}")
    print(f"      📁 파일명: {output_path}")
    print(f"      📊 레코드: {len(tweets)}개")
    print(f"      🖼️  이미지: {sum(1 for t in tweets if t.get('screenshot_path') and os.path.exists(t.get('screenshot_path')))}개")
    print(f"{'─' * 70}\n")

# -----------------------
# CSV 저장 (백업용)
# -----------------------
def save_to_csv(tweets, output_path):
    """CSV 저장 (이미지 없는 버전)"""
    if not tweets:
        print("  ⚠️  저장할 데이터 없음")
        return

    # 요청된 컬럼 순서
    fieldnames = ["등록일자", "게시물번호", "좋아요수", "리트윗수", "댓글수",
                  "내용", "해시태그", "게시물URL", "작성자URL", "작성자아이디"]

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for tweet in tweets:
            mapped_tweet = {
                "등록일자": tweet.get("time_iso_utc", ""),
                "게시물번호": tweet.get("status_id", ""),
                "좋아요수": tweet.get("like_count", "0"),
                "리트윗수": tweet.get("retweet_count", "0"),
                "댓글수": tweet.get("reply_count", "0"),
                "내용": tweet.get("text", ""),
                "해시태그": tweet.get("hashtags", ""),
                "게시물URL": tweet.get("url", ""),
                "작성자URL": tweet.get("url", "").rsplit("/status/", 1)[0] if tweet.get("url") else "",
                "작성자아이디": tweet.get("author_handle", ""),
            }
            writer.writerow(mapped_tweet)

    print(f"\n{'─' * 70}")
    print(f"  💾 CSV 파일 저장 완료!")
    print(f"  {'─' * 66}")
    print(f"      📁 파일명: {output_path}")
    print(f"      📊 레코드: {len(tweets)}개")
    print(f"{'─' * 70}\n")

# -----------------------
# 메인
# -----------------------
def main():
    print(f"\n{'═' * 70}")
    print(f"  🐦 트위터 인용글 추출기 (쿠키 직접 입력 버전)")
    print(f"{'═' * 70}\n")

    # 쿠키 입력받기
    cookies = get_cookies_from_user()
    if not cookies:
        print(f"\n  ❌ 쿠키 입력이 취소되었습니다")
        return

    # .env 파일로 저장할지 물어보기
    print(f"\n{'─' * 70}")
    save_choice = input("  💾 입력한 쿠키를 .env 파일로 저장하시겠습니까? (y/n, 엔터 = n): ").strip().lower()
    if save_choice in ['y', 'yes']:
        save_to_env_file(cookies)

    # 인용글 URL 입력
    print(f"\n{'─' * 70}")
    url = input("  🔗 인용글 URL: ").strip()
    if not url:
        print(f"  ❌ URL을 입력하세요")
        return

    if "/quotes" not in url:
        url = url.rstrip("/") + "/quotes"

    output = input(f"  📁 저장 파일명 (엔터 = quotes.xlsx): ").strip() or "quotes.xlsx"
    # 확장자 자동 처리
    if not output.endswith(".xlsx") and not output.endswith(".csv"):
        output += ".xlsx"

    max_scrolls_input = input(f"  🔢 최대 스크롤 (엔터 = 무제한): ").strip()
    max_scrolls = int(max_scrolls_input) if max_scrolls_input else None

    max_tweets_input = input(f"  🎯 최대 수집 개수 (엔터 = 무제한): ").strip()
    max_tweets = int(max_tweets_input) if max_tweets_input else None

    headless_input = input(f"  👻 브라우저 숨김? (y/n, 엔터 = n): ").strip().lower()
    headless = headless_input in ['y', 'yes']

    # 스크린샷 캡처 여부
    screenshot_input = input(f"  📸 스크린샷 캡처? (y/n, 엔터 = y): ").strip().lower()
    capture_screenshots = screenshot_input != 'n'

    # 봇 회피 기능 안내
    print(f"\n{'═' * 70}")
    print(f"  🛡️  봇 회피 기능 활성화")
    print(f"  {'─' * 66}")
    print(f"      ✅ 점진적 딜레이 증가 (60개 이후 자동)")
    print(f"      ✅ 간헐적 휴식 (60, 120, 180, 240, 300개마다)")
    if USE_UNDETECTED:
        print(f"      ✅ undetected-chromedriver (고급 봇 감지 우회)")
    else:
        print(f"      ⚠️  일반 WebDriver 사용 (권장: pip install undetected-chromedriver)")
    print(f"{'═' * 70}")

    # 실행
    driver = None
    try:
        print(f"\n  🚀 Chrome 드라이버 시작 중...")
        driver = make_driver(headless=headless)

        # 쿠키 로드
        if not load_cookies_to_driver(driver, cookies):
            print(f"\n  ❌ 쿠키 로드 실패")
            return

        # 추출 (봇 회피 기능 포함)
        tweets = extract_quote_tweets(driver, url, max_scrolls=max_scrolls, max_tweets=max_tweets, capture_screenshots=capture_screenshots)

        # 저장 (파일 확장자에 따라 Excel 또는 CSV로 저장)
        if tweets:
            if output.endswith('.xlsx'):
                save_to_excel(tweets, output)
            else:
                save_to_csv(tweets, output)
            print(f"{'═' * 70}")
            print(f"  ✅ 모든 작업이 성공적으로 완료되었습니다!")
            print(f"{'═' * 70}\n")
        else:
            print(f"  ⚠️  추출된 데이터 없음")

    except KeyboardInterrupt:
        print(f"\n\n  ⏹️  사용자에 의해 중단됨")
    except Exception as e:
        print(f"\n  ❌ 오류 발생: {e}")
    finally:
        if driver:
            driver.quit()
            print(f"  🔚 브라우저 종료\n")

if __name__ == "__main__":
    main()
